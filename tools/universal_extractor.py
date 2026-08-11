# tools/universal_extractor.py
"""
Universal file extraction service.
Handles any file type (TXT, CSV, XLSX, PDF, JSON, DOCX, etc.) and extracts
structured or semi-structured data using Grok AI for intelligent parsing.

NOTE: The tool-handler wrapper for this lives in tools/handlers.py
(extract_file_universal), which is the function registered against the
ToolDefinition row and called by ToolService. This module only exposes
the UniversalFileExtractor class — do not add a duplicate module-level
extract_file_universal() function here again.
"""
import logging
import json
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from docx import Document
except ImportError:
    Document = None

logger = logging.getLogger(__name__)


class UniversalFileExtractor:
    """
    Intelligent file extraction that works with any file type.
    Uses Grok AI to understand file structure and extract data intelligently.
    """

    SUPPORTED_FORMATS = {
        '.txt': 'text',
        '.csv': 'spreadsheet',
        '.xlsx': 'spreadsheet',
        '.xls': 'spreadsheet',
        '.pdf': 'document',
        '.json': 'json',
        '.docx': 'document',
        '.doc': 'document',
        '.xml': 'markup',
        '.html': 'markup',
        '.htm': 'markup',
    }

    @staticmethod
    def extract(file_obj, filename: str, context: str = "") -> dict:
        """
        Main entry point for universal file extraction.

        Args:
            file_obj: File-like object opened in binary mode
            filename: Original filename (used to determine type)
            context: Optional domain context (e.g., "invoice", "financial report")

        Returns:
            {
                "success": bool,
                "filename": str,
                "file_type": str,
                "raw_content": str,
                "structured_data": dict,
                "metadata": dict,
                "error": str (if success=False)
            }
        """
        try:
            ext = Path(filename).suffix.lower()
            file_type = UniversalFileExtractor.SUPPORTED_FORMATS.get(ext, 'unknown')

            raw_content = UniversalFileExtractor._extract_raw_content(
                file_obj, filename, ext
            )

            if not raw_content:
                return {
                    "success": False,
                    "filename": filename,
                    "file_type": file_type,
                    "error": f"Could not extract content from {ext} file",
                }

            structured_data = UniversalFileExtractor._parse_with_grok(
                raw_content, filename, context
            )

            return {
                "success": True,
                "filename": filename,
                "file_type": file_type,
                "raw_content": raw_content[:10000],
                "structured_data": structured_data,
                "metadata": {
                    "extension": ext,
                    "file_type": file_type,
                    "content_length": len(raw_content),
                    "context": context,
                },
            }
        except Exception as exc:
            logger.exception(f"Universal extraction failed for {filename}: {exc}")
            return {
                "success": False,
                "filename": filename,
                "error": str(exc),
            }

    @staticmethod
    def _extract_raw_content(file_obj, filename: str, ext: str) -> str:
        """Extract raw text content from any file type."""
        try:
            if ext == '.txt':
                file_obj.seek(0)
                return file_obj.read().decode('utf-8', errors='replace')

            elif ext == '.csv':
                file_obj.seek(0)
                return file_obj.read().decode('utf-8', errors='replace')

            elif ext in ('.xlsx', '.xls'):
                if not openpyxl:
                    return "[XLSX format not supported - openpyxl not installed]"
                file_obj.seek(0)
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    parts.append(f"=== Sheet: {sheet_name} ===")
                    for row in ws.iter_rows(values_only=True):
                        if any(c is not None for c in row):
                            parts.append("\t".join(str(c) if c is not None else "" for c in row))
                return "\n".join(parts)

            elif ext == '.pdf':
                if not pdfplumber:
                    return "[PDF format not supported - pdfplumber not installed]"
                file_obj.seek(0)
                parts = []
                with pdfplumber.open(file_obj) as pdf:
                    for i, page in enumerate(pdf.pages, 1):
                        parts.append(f"\n--- Page {i} ---")
                        text = page.extract_text() or ""
                        parts.append(text)
                        for table in (page.extract_tables() or []):
                            parts.append("[TABLE]")
                            for row in table:
                                parts.append("\t".join(str(c) if c is not None else "" for c in row))
                return "\n".join(parts)

            elif ext == '.json':
                file_obj.seek(0)
                return file_obj.read().decode('utf-8', errors='replace')

            elif ext in ('.docx', '.doc'):
                if not Document:
                    return "[DOCX format not supported - python-docx not installed]"
                file_obj.seek(0)
                doc = Document(file_obj)
                parts = [p.text for p in doc.paragraphs if p.text.strip()]
                return "\n".join(parts)

            elif ext in ('.xml', '.html', '.htm'):
                file_obj.seek(0)
                return file_obj.read().decode('utf-8', errors='replace')

            else:
                file_obj.seek(0)
                return file_obj.read().decode('utf-8', errors='replace')

        except Exception as exc:
            logger.warning(f"Failed to extract raw content from {filename}: {exc}")
            return ""

    @staticmethod
    def _parse_with_grok(raw_content: str, filename: str, context: str = "") -> dict:
        """
        Use Grok AI to intelligently parse the extracted content.
        Returns structured data in JSON format.
        """
        try:
            from ai_engine.services import _get_client, GROK_MODEL
        except ImportError:
            logger.warning("Grok client not available - returning raw content as fallback")
            return {"raw_content": raw_content[:5000]}

        system_prompt = """
You are an expert data extraction specialist. Analyze the provided file content and extract
all structured information into clean JSON format.

Return valid JSON only. Never add explanations outside the JSON.

For any file type, extract:
1. Document type/category
2. Key-value pairs (metadata)
3. Tables/records (as array of objects)
4. Summary statistics
5. Any anomalies or data quality issues

Normalize all data types appropriately (numbers as numbers, dates as YYYY-MM-DD, etc).
"""

        user_prompt = f"""
File: {filename}
Domain Context: {context or 'General data extraction'}

--- FILE CONTENT ---
{raw_content[:8000]}
--- END CONTENT ---

Extract all data into this JSON structure:
{{
  "document_type": "...",
  "metadata": {{}},
  "records": [],
  "summary": {{}},
  "data_quality_issues": []
}}
"""

        try:
            client = _get_client()
            response = client.chat.completions.create(
                model=GROK_MODEL(),
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.0,
                max_tokens=4096,
            )

            raw_response = response.choices[0].message.content
            return UniversalFileExtractor._safe_json_parse(raw_response)

        except Exception as exc:
            logger.warning(f"Grok parsing failed for {filename}: {exc}")
            return {"error": str(exc), "raw_content": raw_content[:5000]}

    @staticmethod
    def _safe_json_parse(text: str) -> dict:
        """Safely parse JSON from LLM output, handling markdown fences and prose."""
        if not text:
            return {"raw_response": ""}
        s = text.strip()

        fence = re.search(r"```(?:json)?\s*(.*?)```", s, re.DOTALL)
        if fence:
            s = fence.group(1).strip()

        try:
            return json.loads(s)
        except json.JSONDecodeError:
            pass

        start, end = s.find("{"), s.rfind("}")
        if start != -1 and end > start:
            try:
                return json.loads(s[start:end + 1])
            except json.JSONDecodeError:
                pass

        logger.warning("Failed to parse JSON response for output starting: %.120s", s)
        return {"raw_response": text[:1000]}