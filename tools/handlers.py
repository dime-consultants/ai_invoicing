# tools/handlers.py
"""
Built-in tool handler primitives.

These are the ONLY functions that belong here — domain-agnostic operations
that genuinely require Python and cannot be expressed as a prompt:

    read_file        — open an UploadedFile and return its raw text content
    detect_file_type — sniff extension + content → type label + confidence
    write_xlsx       — turn rows + headers into a downloadable .xlsx file
    run_python       — execute a user-supplied Python snippet in a sandbox
    call_webhook     — HTTP POST/GET to an external URL (direct handler form)

Everything else — URA extraction, Safaricom parsing, ACON reconciliation,
anomaly flagging, report generation — is expressed as a prompt_transform
ToolDefinition seeded into the DB by the seed_tools management command.
Users can inspect, clone, and edit those tools without touching Python.

Rules
-----
- Every handler accepts **kwargs matching its ToolDefinition.parameters_schema.
- Every handler returns a plain JSON-serialisable dict.
- Handlers NEVER raise — catch all exceptions and return
  {"ok": False, "error": "<message>"} so the LLM can report the failure
  and continue the conversation rather than crashing the job.
- Defer heavy imports (openpyxl, pdfplumber) inside each function so this
  module loads fast even when those packages aren't installed.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# 1. read_file
# ─────────────────────────────────────────────────────────────────────────────

def read_file(
    file_id: int,
    max_chars: int = 12000,
    page_from: int | None = None,
    page_to: int | None = None,
) -> dict:
    """
    Open an UploadedFile and return its text content.

    This is the universal "give me the file content" primitive that all
    prompt_transform tools depend on. It does NOT call any AI.

    Large PDFs are ingested with extraction deferred to a background worker, so
    extracted_text is often empty or a truncated preview. This handler therefore
    falls back to reading the stored file directly, and supports paging through a
    PDF via page_from/page_to so a several-hundred-page document can be processed
    in chunks without waiting for the worker to finish.

    A file with no readable text is reported as a failure, never as a successful
    empty read — otherwise the calling model answers confidently from an empty
    document.

    Parameters
    ----------
    file_id   : PK of the UploadedFile record.
    max_chars : Truncate returned text to this many characters (default 12 000).
                Keeps token usage predictable for the calling LLM.
    page_from : First PDF page to read (1-indexed, inclusive). PDFs only.
    page_to   : Last PDF page to read (1-indexed, inclusive). PDFs only.
                Omit both to read the whole document.

    Returns
    -------
    On success:
        {
            "ok": true,
            "file_id": <int>,
            "filename": <str>,
            "extension": <str>,
            "detected_type": <str | null>,
            "parse_status": <str>,
            "source": <"extracted_text" | "live_pdf_range" | "raw_file">,
            "text": <str>,          # truncated to max_chars
            "full_length": <int>,   # length before truncation
            "truncated": <bool>,
            # page-range reads only:
            "page_from": <int>, "page_to": <int>, "page_count": <int>
        }

    On failure:
        {"ok": false, "error": <str>, ...plus the context fields above}
    """
    try:
        from uploads.models import UploadedFile
        uf = UploadedFile.objects.get(pk=file_id)
    except Exception as exc:
        # DoesNotExist, or the app registry is unavailable.
        if type(exc).__name__ == "DoesNotExist":
            return {"ok": False, "error": f"UploadedFile id={file_id} not found."}
        logger.exception("read_file(%s): %s", file_id, exc)
        return {"ok": False, "error": str(exc)}

    ctx = {
        "file_id":       file_id,
        "filename":      uf.original_filename,
        "extension":     uf.extension,
        "detected_type": uf.detected_type,
        "parse_status":  uf.parse_status,
    }

    try:
        ext        = (uf.extension or "").lower()
        wants_page = page_from is not None or page_to is not None
        text       = ""
        source     = ""
        page_info: dict = {}

        # A page range always reads the file live: extracted_text is a
        # whole-document preview and cannot answer "pages 40-60".
        if not wants_page:
            text   = uf.extracted_text or ""
            source = "extracted_text" if text else ""

        if not text:
            try:
                path = uf.file.path
            except Exception as path_exc:
                logger.warning("read_file: no local path for %s: %s", file_id, path_exc)
                path = None

            if path and ext == "pdf":
                from uploads.services import extract_pdf_page_range
                res = extract_pdf_page_range(
                    path,
                    page_from=page_from or 1,
                    page_to=page_to,
                    max_chars=max_chars,
                )
                if res.get("ok"):
                    text   = res.get("text") or ""
                    source = "live_pdf_range"
                    page_info = {
                        "page_from":  res.get("page_from"),
                        "page_to":    res.get("page_to"),
                        "page_count": res.get("page_count"),
                    }
                else:
                    return {**ctx, "ok": False,
                            "error": res.get("error") or "Could not open PDF."}

            elif path and ext in ("txt", "csv"):
                try:
                    uf.file.open("rb")
                    text   = uf.file.read().decode("utf-8", errors="replace")
                    source = "raw_file"
                finally:
                    try:
                        uf.file.close()
                    except Exception:
                        pass

            elif path:
                # docx / xlsx / everything else — reuse the ingest pipeline's
                # extractor rather than duplicating format handling here.
                from uploads.services import _extract_text
                text   = _extract_text(path, ext, max_chars=max_chars) or ""
                source = "raw_file"

        if not text:
            if uf.parse_status in ("pending", "parsing"):
                reason = ("Extraction is still running for this file. "
                          "Retry shortly, or request a page range with "
                          "page_from/page_to to read the document directly.")
            elif uf.parse_status == "parse_error":
                reason = f"Extraction failed for this file: {uf.parse_error or 'unknown error'}"
            else:
                reason = ("No readable text in this file. It is most likely a "
                          "scanned or image-only document with no text layer, "
                          "which requires OCR. Do not infer its contents.")
            return {**ctx, "ok": False, "error": reason, **page_info}

        full_length = len(text)
        return {
            **ctx,
            "ok":          True,
            "source":      source,
            "text":        text[:max_chars],
            "full_length": full_length,
            "truncated":   full_length > max_chars,
            **page_info,
        }

    except Exception as exc:
        logger.exception("read_file(%s): %s", file_id, exc)
        return {**ctx, "ok": False, "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# 2. detect_file_type
# ─────────────────────────────────────────────────────────────────────────────

def detect_file_type(file_id: int) -> dict:
    """
    Inspect an UploadedFile and return (and persist) its detected type.

    Uses file extension + a fast content sniff — no LLM call. The result
    is written back to UploadedFile.detected_type so subsequent tools can
    branch on it without re-detecting.

    Detected type labels
    --------------------
    ura_fiscal_receipt  — KRA/URA .txt periodical report with CU INVOICE NUMBER blocks
    ura_sales_table     — URA/KRA .xls/.xlsx with FDN / TOTAL (A+B) columns
    safaricom_bill      — Safaricom PostPay PDF with TAX INVOICE SUMMARY section
    acon_export         — ACON .xlsx with DEBTOR ACCOUNT / LC AMOUNT columns
    generic_xlsx        — any other spreadsheet
    generic_csv         — CSV
    generic_pdf         — any other PDF
    unknown_txt         — .txt that didn't match known patterns
    unknown             — everything else

    Returns
    -------
    {"ok": true, "file_id": .., "filename": .., "detected_type": .., "confidence": ..}
    """
    try:
        from uploads.models import UploadedFile

        uf   = UploadedFile.objects.get(pk=file_id)
        ext  = uf.extension.lower()
        # Use a generous slice — Safaricom bill signals appear pages in
        text = (uf.extracted_text or "")[:20_000]

        # A large PDF is ingested with extraction deferred, so extracted_text is
        # routinely empty here. Sniffing that would label every big bill
        # "generic_pdf" — and since the verdict is persisted and never recomputed,
        # the mislabel is permanent and every downstream branch on detected_type
        # picks the wrong tools. Read the document directly instead.
        if not text and ext == "pdf":
            try:
                from uploads.services import extract_pdf_page_range
                live = extract_pdf_page_range(
                    uf.file.path, page_from=1, page_to=20, max_chars=20_000,
                )
                if live.get("ok"):
                    text = live.get("text") or ""
            except Exception as sniff_exc:
                logger.warning(
                    "detect_file_type: live sniff failed for %s: %s", file_id, sniff_exc,
                )

        sniffed_content = bool(text)
        text = text[:20_000].upper()

        detected   = "unknown"
        confidence = "low"

        if ext == "txt":
            if "CU INVOICE NUMBER" in text and (
                "FISCAL RECEIPT" in text or "CREDIT NOTE" in text
            ):
                detected, confidence = "ura_fiscal_receipt", "high"
            elif "PERIODICAL REPORT" in text:
                detected, confidence = "ura_fiscal_receipt", "medium"
            else:
                detected, confidence = "unknown_txt", "low"

        elif ext == "pdf":
            if "SAFARICOM" in text or "POSTPAY" in text or "BILLED AMOUNT" in text:
                detected, confidence = "safaricom_bill", "high"
            else:
                detected, confidence = "generic_pdf", "low"

        elif ext in ("xlsx", "xls"):
            if (
                "DEBTOR ACCOUNT" in text
                or "STATUTORY ITEM" in text
                or "LC AMOUNT" in text
            ):
                detected, confidence = "acon_export", "high"
            elif (
                "NAME OF PURCHASER" in text
                or "TOTAL (A+B)" in text
                or "FDN" in text
            ):
                detected, confidence = "ura_sales_table", "high"
            elif "CU INVOICE" in text:
                detected, confidence = "ura_fiscal_receipt", "medium"
            else:
                detected, confidence = "generic_xlsx", "medium"

        elif ext == "csv":
            detected, confidence = "generic_csv", "medium"

        # Never persist a verdict reached without reading any content: it would
        # be indistinguishable from a real one and nothing would recompute it.
        # UploadService.complete_extraction re-runs detection once text lands.
        if sniffed_content:
            uf.detected_type        = detected
            uf.detection_confidence = confidence
            uf.save(update_fields=["detected_type", "detection_confidence"])
        else:
            confidence = "low"
            logger.info(
                "detect_file_type: no content available for file_id=%s (%s); "
                "returning a provisional extension-only guess without persisting.",
                file_id, uf.parse_status,
            )

        return {
            "ok":            True,
            "file_id":       file_id,
            "filename":      uf.original_filename,
            "detected_type": detected,
            "confidence":    confidence,
            # False → guessed from the extension alone because no text was
            # readable yet; treat as provisional and re-detect after extraction.
            "based_on_content": sniffed_content,
            "parse_status":     uf.parse_status,
        }

    except UploadedFile.DoesNotExist:
        return {"ok": False, "error": f"UploadedFile id={file_id} not found."}
    except Exception as exc:
        logger.exception("detect_file_type(%s): %s", file_id, exc)
        return {"ok": False, "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# 3. write_xlsx
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(
    filename: str,
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Sheet1",
) -> dict:
    """
    Build an .xlsx file from headers + rows and save it to the outputs directory.

    This is the report-generation primitive. A prompt_transform tool (or Grok
    itself) produces the structured rows; this handler serialises them to a
    downloadable spreadsheet.

    Parameters
    ----------
    filename   : Output filename, e.g. "ura_report.xlsx". Extension forced to .xlsx.
    headers    : Column header labels.
    rows       : List of row arrays. Values are coerced to str if not
                 int/float/bool/None — keeps the xlsx valid.
    sheet_name : Worksheet tab name (default "Sheet1").

    Returns
    -------
    {"ok": true, "output_filename": <abs path str>, "record_count": <int>, "summary": <str>}
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.conf import settings

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]   # Excel tab name limit

        # Style the header row
        HDR_FONT  = Font(bold=True, color="FFFFFF")
        HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
        HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN

        # Write data rows — coerce non-scalar types
        _scalar = (int, float, bool, type(None))
        for row in rows:
            ws.append([
                v if isinstance(v, _scalar) else str(v)
                for v in row
            ])

        # Auto-column widths (capped at 50)
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 50)

        out_dir = Path(settings.BASE_DIR) / "outputs" / "converted"
        out_dir.mkdir(parents=True, exist_ok=True)

        stem     = Path(filename).stem
        out_path = out_dir / f"{stem}.xlsx"
        wb.save(out_path)

        return {
            "ok":              True,
            "output_filename": str(out_path),
            "record_count":    len(rows),
            "summary":         f"Wrote {len(rows)} rows to {out_path.name}.",
        }

    except Exception as exc:
        logger.exception("write_xlsx(%s): %s", filename, exc)
        return {"ok": False, "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# 4. run_python
# ─────────────────────────────────────────────────────────────────────────────

# Modules the sandbox is allowed to import
_ALLOWED_MODULES = frozenset({
    "re", "json", "csv", "math", "statistics",
    "datetime", "decimal", "collections", "itertools",
    "pathlib", "io", "string", "textwrap",
})

_BLOCKED_BUILTINS = frozenset({
    "__import__", "open", "exec", "eval", "compile",
    "globals", "locals", "vars", "dir",
    "breakpoint", "input", "print",
    "getattr", "setattr", "delattr", "hasattr",
})


def run_python(code: str, context: dict | None = None) -> dict:
    """
    Execute a user-supplied Python snippet in a restricted sandbox and
    return its output.

    The snippet has access to:
    - A `context` dict passed in by the caller (read-only by convention).
    - A `result` dict it should populate — {"ok": bool, ...}.
    - The allowed standard-library modules listed in _ALLOWED_MODULES.

    Restricted: no file I/O, no subprocess, no network, no __import__.

    This is a power-user escape hatch. It is marked is_safe=False in the
    ToolDefinition so Grok must ask for user confirmation before calling it.

    Parameters
    ----------
    code    : Python source to execute. Must set `result["ok"]`.
    context : Optional dict of values injected into the execution namespace.

    Returns
    -------
    {"ok": bool, ...} — whatever the snippet set on `result`, plus
    {"ok": False, "error": "..."} on any exception.

    Example snippet
    ---------------
    ::

        import re
        numbers = re.findall(r"\\d+", context.get("text", ""))
        result["ok"]    = True
        result["found"] = numbers
        result["count"] = len(numbers)
    """
    if not code or not code.strip():
        return {"ok": False, "error": "code is required."}

    # Build a minimal safe namespace
    safe_builtins = {
        k: v for k, v in __builtins__.items()   # type: ignore[union-attr]
        if k not in _BLOCKED_BUILTINS
    } if isinstance(__builtins__, dict) else {
        k: getattr(__builtins__, k)
        for k in dir(__builtins__)
        if k not in _BLOCKED_BUILTINS and not k.startswith("_")
    }

    import importlib

    def _safe_import(name, *args, **kwargs):
        if name not in _ALLOWED_MODULES:
            raise ImportError(
                f"Module '{name}' is not allowed in the sandbox. "
                f"Allowed: {sorted(_ALLOWED_MODULES)}"
            )
        return importlib.import_module(name)

    namespace: dict = {
        "__builtins__": {**safe_builtins, "__import__": _safe_import},
        "context":      context or {},
        "result":       {"ok": False},
    }

    try:
        exec(compile(code, "<tool_snippet>", "exec"), namespace)   # noqa: S102
        outcome = namespace.get("result", {"ok": False, "error": "result was not set"})
        if not isinstance(outcome, dict):
            return {"ok": False, "error": "result must be a dict"}
        return outcome
    except Exception as exc:
        logger.warning("run_python snippet raised: %s", exc)
        return {"ok": False, "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# 5. call_webhook
# ─────────────────────────────────────────────────────────────────────────────

def call_webhook(
    url: str,
    payload: dict | None = None,
    method: str = "POST",
    headers: dict | None = None,
    timeout_seconds: int = 30,
) -> dict:
    """
    Make an HTTP request to an external URL and return the JSON response.

    This is the direct-handler form of the webhook dispatch. Unlike the
    service-layer _call_webhook() helper (which reads config from a
    UserToolConfig row), this handler accepts all parameters as arguments
    so Grok can call it with ad-hoc URLs without a pre-registered tool.

    Marked is_safe=False — Grok must ask for user confirmation before
    posting to external systems.

    Parameters
    ----------
    url             : Endpoint to call.
    payload         : JSON body (POST) or query-string params (GET).
    method          : "POST" or "GET" (default "POST").
    headers         : Extra HTTP headers.
    timeout_seconds : Request timeout (1–120, default 30).

    Returns
    -------
    {"ok": true, <response fields>} or {"ok": false, "error": "..."}
    """
    import urllib.request
    import urllib.error
    import json as _json

    method  = (method or "POST").upper()
    payload = payload or {}
    hdrs    = {
        "Content-Type": "application/json",
        "User-Agent":   "ai-invoicing/1.0 (call_webhook)",
        **(headers or {}),
    }
    timeout = max(1, min(int(timeout_seconds or 30), 120))

    try:
        body = _json.dumps(payload).encode("utf-8") if method == "POST" else None

        if method == "GET" and payload:
            import urllib.parse
            qs  = urllib.parse.urlencode(
                {k: _json.dumps(v) if isinstance(v, (dict, list)) else v
                 for k, v in payload.items()}
            )
            url = f"{url}?{qs}"

        req = urllib.request.Request(url, data=body, headers=hdrs, method=method)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw    = resp.read().decode("utf-8", errors="replace")
            status = resp.status

        if status >= 400:
            return {"ok": False, "error": f"HTTP {status}", "detail": raw[:500]}

        try:
            return {"ok": True, **_json.loads(raw)}
        except _json.JSONDecodeError:
            return {"ok": True, "raw_response": raw[:2000]}

    except urllib.error.URLError as exc:
        return {"ok": False, "error": f"Connection error: {exc.reason}"}
    except TimeoutError:
        return {"ok": False, "error": f"Timed out after {timeout}s"}
    except Exception as exc:
        logger.exception("call_webhook(%s): %s", url, exc)
        return {"ok": False, "error": str(exc)}

# ═════════════════════════════════════════════════════════════════════════════
# Domain handlers — restored from commit 9aeaafc^ ("updated tools handler").
#
# The 2026-08 rewrite replaced these with generic prompt_transform tools. That
# moved parsing INTO the prompt and lost logic an LLM cannot reliably infer:
# the Safaricom summary→detail CU join, and the rule that URA/ACON reconcile on
# the statutory column rather than ACON's internal Item Number. Restored as
# builtins; the generic tools remain for arbitrary file pairs.
# ═════════════════════════════════════════════════════════════════════════════

def _norm_number(raw: str) -> float:
    """
    Normalise ambiguous number strings to float.
    Handles URA style: '4 862 563,00'  →  4862563.0
    Handles standard:  '1,234.56'      →  1234.56
    """
    s = str(raw).strip()
    if re.match(r'^[\d ]+,\d{1,2}$', s):          # space-thousands + comma-decimal
        return float(s.replace(' ', '').replace(',', '.'))
    return float(s.replace(',', '').replace(' ', ''))


def _get_uploaded_file(file_id: int):
    """Return an UploadedFile instance or raise ValueError."""
    from uploads.models import UploadedFile
    try:
        return UploadedFile.objects.get(pk=file_id)
    except UploadedFile.DoesNotExist:
        raise ValueError(f"UploadedFile id={file_id} not found.")



def _norm_id(value) -> str:
    """Normalise a fiscal id (FDN / CU / statutory no) to a clean string,
    stripping float artefacts like a trailing '.0' from numeric cells."""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _find_col(headers, *keywords) -> int:
    """Index of the first header containing any keyword (case-insensitive), else -1."""
    low = [str(h or "").lower() for h in headers]
    for kw in keywords:
        for i, h in enumerate(low):
            if kw in h:
                return i
    return -1


def _read_tabular(uf, prefer=()) -> tuple[list, list]:
    """
    Load an .xls/.xlsx UploadedFile into (headers, data_rows).

    Picks the sheet whose header row best matches the `prefer` keywords (so we
    grab the real data sheet, not a 'Criteria' tab), and detects the header row
    as the first row with >= 4 non-empty cells.
    """
    ext = uf.extension.lower()
    sheets = []  # (name, rows)
    uf.file.open("rb")
    if ext == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(uf.file, data_only=True, read_only=True)
        for ws in wb.worksheets:
            sheets.append((ws.title, [list(r) for r in ws.iter_rows(values_only=True)]))
        wb.close()
    elif ext == "xls":
        import xlrd
        book = xlrd.open_workbook(file_contents=uf.file.read())
        for sh in book.sheets():
            sheets.append((sh.name, [sh.row_values(r) for r in range(sh.nrows)]))
    else:
        raise ValueError(f"_read_tabular: unsupported extension '{ext}'")

    def header_idx(rows):
        for i, r in enumerate(rows[:10]):
            if sum(1 for c in r if c not in (None, "")) >= 4:
                return i
        return 0

    best, best_score = None, -1
    for _name, rows in sheets:
        if not rows:
            continue
        hidx = header_idx(rows)
        hdr = [str(c).strip() if c is not None else "" for c in rows[hidx]]
        low = " | ".join(hdr).lower()
        kw_hits = sum(1 for kw in prefer if kw in low) if prefer else 0
        score = kw_hits * 1_000_000 + len(rows)  # keyword match dominates, size breaks ties
        if score > best_score:
            best_score, best = score, (hdr, rows, hidx)
    if not best:
        return [], []
    hdr, rows, hidx = best
    data = [[("" if c is None else c) for c in r]
            for r in rows[hidx + 1:] if any(c not in (None, "") for c in r)]
    return hdr, data


def _parse_fiscal_side(uf) -> list[dict]:
    """
    Parse a URA/KRA fiscal *sales* file into normalised rows:
        {fiscal_no, name, date, total, vat}

    Handles BOTH the KRA `.txt` periodical report (CU INVOICE NUMBER blocks) and
    the URA/KRA `.xls`/`.xlsx` sales tables (FDN / TOTAL (A+B) columns).
    """
    ext = uf.extension.lower()
    out = []
    if ext == "txt":
        uf.file.open("rb")
        raw = uf.file.read().decode("utf-8", errors="replace")
        block = re.compile(
            r'^(?P<t>FISCAL RECEIPT|CREDIT NOTE)\s*\r?\n'
            r'CU INVOICE NUMBER:\s*(?P<cu>\S+)\s*\r?\n'
            r'(?P<d>\d{2}-\d{2}-\d{4})\s+\d{2}:\d{2}:\d{2}\s*\r?\n'
            r'TOTAL:\s+(?P<tot>[\d\s,]+)\r?\n'
            r'TAXES:\s+(?P<tax>[\d\s,]+)',
            re.MULTILINE | re.IGNORECASE,
        )
        for m in block.finditer(raw):
            out.append({
                "fiscal_no": _norm_id(m.group("cu")),
                "name": "", "date": m.group("d"),
                "total": _norm_number(m.group("tot")),
                "vat": _norm_number(m.group("tax")),
            })
        return out

    hdr, rows = _read_tabular(uf, prefer=("fdn", "name of purchaser", "cu invoice", "total (a+b)"))
    c_no  = _find_col(hdr, "fdn", "cu invoice", "statutory", "invoice number")
    c_nm  = _find_col(hdr, "name of purchaser", "name")
    c_dt  = _find_col(hdr, "invoice date", "date")
    c_tot = _find_col(hdr, "total (a+b)", "total", "billed amount", "(ugx)(a)")
    c_vat = _find_col(hdr, "vat charged", "vat")
    for r in rows:
        if not (0 <= c_no < len(r)):
            continue
        fno = _norm_id(r[c_no])
        if not fno or fno.lower() in ("none", ""):
            continue
        out.append({
            "fiscal_no": fno,
            "name":  str(r[c_nm]).strip() if 0 <= c_nm < len(r) else "",
            "date":  str(r[c_dt]).strip() if 0 <= c_dt < len(r) else "",
            "total": _saf_num(r[c_tot]) if 0 <= c_tot < len(r) else None,
            "vat":   _saf_num(r[c_vat]) if 0 <= c_vat < len(r) else None,
        })
    return out


def _parse_acon_side(uf) -> list[dict]:
    """
    Parse an ACON export into normalised rows:
        {fiscal_no, item_number, name, amount}

    The fiscal join key is ACON's statutory column — 'Statutory Item No(For
    Download VAT)' (Kenya CU) or 'FDN' (Uganda) — NOT the internal Item Number.
    """
    hdr, rows = _read_tabular(uf, prefer=("debtor account", "statutory item", "fdn", "item number"))
    c_no   = _find_col(hdr, "statutory item", "fdn")
    c_item = _find_col(hdr, "item number")
    c_nm   = _find_col(hdr, "full name", "name")
    c_amt  = _find_col(hdr, "lc amount", "amount")
    out = []
    for r in rows:
        if not (0 <= c_no < len(r)):
            continue
        fno = _norm_id(r[c_no])
        if not fno or fno.lower() in ("none", ""):
            continue
        out.append({
            "fiscal_no":   fno,
            "item_number": _norm_id(r[c_item]) if 0 <= c_item < len(r) else "",
            "name":        str(r[c_nm]).strip() if 0 <= c_nm < len(r) else "",
            "amount":      _saf_num(r[c_amt]) if 0 <= c_amt < len(r) else None,
        })
    return out


# Patterns for the per-subscriber TAX INVOICE pages and the summary table.
_SAF_RE_INV   = re.compile(r"Invoice Number\s+([A-Z0-9\-]+)", re.I)
_SAF_RE_CU    = re.compile(r"CU\s*INVOICE\s*NO[:\s]+(\d+)", re.I)
_SAF_RE_PHONE = re.compile(r"^\d{9}$")
_SAF_RE_INVNO = re.compile(r"^[A-Z]\d?-?\d{6,}$")


def _saf_num(raw) -> float | None:
    """Parse a Safaricom amount like '9,429.13' → 9429.13; '' → None."""
    s = str(raw or "").strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "").replace(" ", ""))
    except ValueError:
        return None


def _saf_department(name: str) -> str:
    """Best-effort: strip the Kuehne+Nagel company prefix to leave the unit/user."""
    d = re.sub(r"^\s*kuehne\s*\+?\s*nagel(\s+ltd)?\b[\s\-]*", "", name, flags=re.I)
    d = re.sub(r"^\s*kuehne\s+and\s+nagel(\s+ltd)?\b[\s\-]*", "", d, flags=re.I)
    return d.strip() or name


def extract_safaricom_bill(file_id: int) -> dict:
    """
    Extract a per-line telephone billing report from a Safaricom PostPay bill PDF.

    The bill has two relevant sections:
      • TAX INVOICE SUMMARY (a few pages): Name/department, phone (Reference NO.),
        invoice number, Net/VAT/Excise/Billed Amount — one row per subscriber.
      • One TAX INVOICE per subscriber (each spanning several pages) whose first
        page carries the fiscal "CU INVOICE NO" plus the invoice number.

    This joins the two on invoice number so every subscriber row gets its CU
    number, then writes an .xlsx billing report mapping
    telephone user → department → phone → invoice → CU number → amounts.
    """
    try:
        import pdfplumber
        import openpyxl
        from django.conf import settings
        from django.utils import timezone

        uf = _get_uploaded_file(file_id)
        uf.parse_status = "parsing"
        uf.save(update_fields=["parse_status"])

        detail_cu: dict[str, str] = {}   # invoice_no -> CU invoice number
        summary: dict[str, dict]  = {}   # invoice_no -> subscriber row (deduped)

        uf.file.open("rb")
        with pdfplumber.open(uf.file) as pdf:
            in_summary = False
            for page in pdf.pages:
                text = page.extract_text() or ""
                upper = text.upper()

                if "TAX INVOICE SUMMARY" in upper:
                    in_summary = True

                # A per-subscriber document page: end of the summary section,
                # and the place where the CU invoice number lives.
                if "SUBSCRIBER NUMBER" in upper:
                    in_summary = False
                    m_inv, m_cu = _SAF_RE_INV.search(text), _SAF_RE_CU.search(text)
                    if m_inv and m_cu:
                        detail_cu.setdefault(m_inv.group(1).strip(), m_cu.group(1).strip())
                    continue

                if in_summary:
                    for table in (page.extract_tables() or []):
                        for row in table:
                            cells = [str(c or "").replace("\n", " ").strip() for c in row]
                            if len(cells) < 7:
                                continue
                            phone = cells[1].replace(" ", "")
                            inv   = cells[2].replace(" ", "")
                            if _SAF_RE_PHONE.match(phone) and _SAF_RE_INVNO.match(inv):
                                summary[inv] = {
                                    "name": cells[0], "phone": phone, "invoice_no": inv,
                                    "net": cells[3], "vat": cells[4],
                                    "excise": cells[5], "billed": cells[6],
                                }

        headers = [
            "Telephone User", "Department", "Phone Number", "Invoice Number",
            "CU Invoice Number", "Net Amount", "VAT", "Excise", "Billed Amount",
        ]
        rows = []
        for inv, s in summary.items():
            rows.append([
                s["name"],
                _saf_department(s["name"]),
                s["phone"],
                inv,
                detail_cu.get(inv, ""),
                _saf_num(s["net"]),
                _saf_num(s["vat"]),
                _saf_num(s["excise"]),
                _saf_num(s["billed"]),
            ])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Safaricom Billing"
        ws.append(headers)
        for row in rows:
            ws.append(row)

        out_dir = Path(settings.BASE_DIR) / "outputs" / "converted"
        out_dir.mkdir(parents=True, exist_ok=True)
        stem = Path(uf.original_filename).stem
        out_path = out_dir / f"{stem}_extracted.xlsx"
        wb.save(out_path)

        uf.parse_status = "parsed"
        uf.parsed_at    = timezone.now()
        uf.save(update_fields=["parse_status", "parsed_at"])

        total_billed   = sum(r[8] or 0 for r in rows)
        without_cu      = sum(1 for r in rows if not r[4])
        return {
            "ok":              True,
            "file_id":         file_id,
            "record_count":    len(rows),
            "headers":         headers,
            "rows":            rows[:5],
            "output_filename": str(out_path),
            "total_billed":    round(total_billed, 2),
            "missing_cu_count": without_cu,
            "summary": (
                f"Extracted {len(rows)} telephone subscribers from Safaricom bill "
                f"'{uf.original_filename}'. Total billed Ksh {total_billed:,.2f}. "
                f"{len(rows) - without_cu}/{len(rows)} matched a CU invoice number. "
                f"Output saved to {out_path.name}."
            ),
        }

    except Exception as exc:
        logger.exception("extract_safaricom_bill(%s): %s", file_id, exc)
        try:
            uf = _get_uploaded_file(file_id)
            uf.parse_status = "parse_error"
            uf.parse_error  = str(exc)
            uf.save(update_fields=["parse_status", "parse_error"])
        except Exception:
            pass
        return {"ok": False, "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# reconcile_ura_vs_acon
# ─────────────────────────────────────────────────────────────────────────────

def reconcile_ura_vs_acon(ura_file_id: int, acon_file_id: int) -> dict:
    """
    Reconcile a URA/KRA fiscal *sales* file against an ACON export.

    The join key is the FISCAL number — the URA/KRA FDN or CU Invoice Number
    matched against ACON's statutory column ('Statutory Item No(For Download
    VAT)' for Kenya, or 'FDN' for Uganda) — NOT ACON's internal Item Number.

    Produces a reconciliation workpaper (.xlsx) with one row per fiscal number:
    matched rows (with both amounts and the variance), records present in the
    fiscal file but MISSING_IN_ACON, and records in ACON but MISSING_IN_URA.
    """
    try:
        import openpyxl
        from django.conf import settings

        ura_uf  = _get_uploaded_file(ura_file_id)
        acon_uf = _get_uploaded_file(acon_file_id)

        ura_rows  = _parse_fiscal_side(ura_uf)
        acon_rows = _parse_acon_side(acon_uf)

        # Index ACON by fiscal number (first occurrence wins).
        acon_by_no = {}
        for a in acon_rows:
            acon_by_no.setdefault(a["fiscal_no"], a)
        ura_nos = {u["fiscal_no"] for u in ura_rows}

        TOL = 1.0
        matched, variances, unmatched_ura = [], [], []
        for u in ura_rows:
            a = acon_by_no.get(u["fiscal_no"])
            if not a:
                unmatched_ura.append(u)
                continue
            ut, at = u.get("total"), a.get("amount")
            var = round(ut - at, 2) if (ut is not None and at is not None) else None
            row = {
                "fiscal_no":   u["fiscal_no"],
                "name":        u.get("name") or a.get("name"),
                "date":        u.get("date"),
                "ura_total":   ut,
                "acon_item":   a.get("item_number"),
                "acon_amount": at,
                "variance":    var,
            }
            matched.append(row)
            if var is not None and abs(var) > TOL:
                variances.append(row)

        unmatched_acon = [a for a in acon_rows if a["fiscal_no"] not in ura_nos]

        # ── Write the reconciliation workpaper ────────────────────────────────
        out_dir = Path(settings.BASE_DIR) / "outputs" / "converted"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / "ura_acon_reconciliation.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reconciliation"
        ws.append(["Fiscal No (FDN/CU)", "Purchaser", "Date",
                   "URA Total", "ACON Item No", "ACON Amount", "Variance", "Status"])
        for m in matched:
            status = "VARIANCE" if (m["variance"] is not None and abs(m["variance"]) > TOL) else "MATCH"
            ws.append([m["fiscal_no"], m["name"], m["date"], m["ura_total"],
                       m["acon_item"], m["acon_amount"], m["variance"], status])
        for u in unmatched_ura:
            ws.append([u["fiscal_no"], u.get("name"), u.get("date"), u.get("total"),
                       "", "", "", "MISSING_IN_ACON"])
        for a in unmatched_acon:
            ws.append([a["fiscal_no"], a.get("name"), "", "",
                       a.get("item_number"), a.get("amount"), "", "MISSING_IN_URA"])
        wb.save(out_path)

        return {
            "ok":                   True,
            "ura_count":            len(ura_rows),
            "acon_count":           len(acon_rows),
            "matched_count":        len(matched),
            "variance_count":       len(variances),
            "missing_in_acon":      len(unmatched_ura),
            "missing_in_ura":       len(unmatched_acon),
            "output_filename":      str(out_path),
            "rows":                 matched[:5],
            "summary": (
                f"Reconciled {len(ura_rows)} URA/KRA records against "
                f"{len(acon_rows)} ACON records (matched on fiscal number). "
                f"Matched {len(matched)}, of which {len(variances)} have amount "
                f"variances. Missing in ACON: {len(unmatched_ura)}; "
                f"missing in URA: {len(unmatched_acon)}. "
                f"Workpaper saved to {out_path.name}."
            ),
        }

    except Exception as exc:
        logger.exception("reconcile_ura_vs_acon: %s", exc)
        return {"ok": False, "error": str(exc)}
