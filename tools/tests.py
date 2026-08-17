"""
Regression tests for the read_file tool handler.

read_file used to return {"ok": True, "text": ""} whenever extracted_text was
empty — which is the normal state for a large PDF whose extraction is deferred
to a background worker. The agent is instructed to call read_file immediately
after upload, so it received a *successful* empty read and answered from an
empty document. It also never implemented the page_from/page_to chunked access
that the whole deferred-PDF design documents and depends on.
"""

import json
import re
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from organizations.models import Organization
from tools.handlers import detect_file_type, export_file, read_file
from tools.models import ToolCall, ToolDefinition
from tools.services import _call_prompt_transform
from uploads.models import UploadedFile
from uploads.services import UploadService


class ReadFileTests(TestCase):

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="tools@example.com", password="pw12345!",
            first_name="T", last_name="U",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    def _file(self, *, name="doc.pdf", content=b"%PDF-1.4", **kwargs):
        upload = SimpleUploadedFile(name, content)
        defaults = dict(
            batch=self.batch, file=upload, original_filename=name,
            file_size_bytes=len(content), extension=name.rsplit(".", 1)[-1],
            parse_status="parsed", extracted_text="",
        )
        defaults.update(kwargs)
        return UploadedFile.objects.create(**defaults)

    # ── the core regression ───────────────────────────────────────────────────

    def test_missing_file_is_reported(self):
        res = read_file(file_id=999999)
        self.assertFalse(res["ok"])
        self.assertIn("not found", res["error"])

    def test_no_text_layer_is_a_failure_not_an_empty_success(self):
        uf = self._file()
        with patch("uploads.services.extract_pdf_page_range",
                   return_value={"ok": True, "text": "", "page_from": 1,
                                 "page_to": 1, "page_count": 1}):
            res = read_file(file_id=uf.pk)
        self.assertFalse(res["ok"], "empty read reported as success")
        self.assertIn("scanned", res["error"].lower())
        self.assertEqual(res["parse_status"], "parsed")

    def test_still_extracting_says_so(self):
        uf = self._file(parse_status="pending", extraction_deferred=True)
        with patch("uploads.services.extract_pdf_page_range",
                   return_value={"ok": True, "text": "", "page_from": 1,
                                 "page_to": 1, "page_count": 1}):
            res = read_file(file_id=uf.pk)
        self.assertFalse(res["ok"])
        self.assertIn("still running", res["error"])

    def test_parse_error_is_surfaced(self):
        uf = self._file(parse_status="parse_error", parse_error="pdfplumber blew up")
        with patch("uploads.services.extract_pdf_page_range",
                   return_value={"ok": False, "error": "Cannot open PDF: bad header"}):
            res = read_file(file_id=uf.pk)
        self.assertFalse(res["ok"])
        self.assertIn("Cannot open PDF", res["error"])

    # ── deferred large PDFs are readable live ─────────────────────────────────

    def test_deferred_pdf_falls_back_to_live_extract(self):
        uf = self._file(parse_status="pending", extraction_deferred=True, page_count=300)
        with patch("uploads.services.extract_pdf_page_range",
                   return_value={"ok": True, "text": "CU 0012345678",
                                 "page_from": 1, "page_to": 300, "page_count": 300}):
            res = read_file(file_id=uf.pk)
        self.assertTrue(res["ok"], "deferred PDF unreadable until the worker finishes")
        self.assertEqual(res["text"], "CU 0012345678")
        self.assertEqual(res["source"], "live_pdf_range")
        self.assertEqual(res["page_count"], 300)

    def test_page_range_is_passed_through_and_bypasses_the_cached_preview(self):
        uf = self._file(extracted_text="whole-document preview", page_count=300)
        with patch("uploads.services.extract_pdf_page_range") as ranged:
            ranged.return_value = {"ok": True, "text": "page 40-60 text",
                                   "page_from": 40, "page_to": 60, "page_count": 300}
            res = read_file(file_id=uf.pk, page_from=40, page_to=60)

        self.assertTrue(res["ok"])
        self.assertEqual(res["text"], "page 40-60 text")
        self.assertEqual((res["page_from"], res["page_to"]), (40, 60))
        kwargs = ranged.call_args.kwargs
        self.assertEqual((kwargs["page_from"], kwargs["page_to"]), (40, 60))

    # ── existing behaviour preserved ──────────────────────────────────────────

    def test_cached_text_is_used_when_present(self):
        uf = self._file(extracted_text="already extracted")
        res = read_file(file_id=uf.pk)
        self.assertTrue(res["ok"])
        self.assertEqual(res["text"], "already extracted")
        self.assertEqual(res["source"], "extracted_text")
        self.assertFalse(res["truncated"])

    def test_txt_raw_fallback_still_works(self):
        uf = self._file(name="notes.txt", content=b"plain text body")
        res = read_file(file_id=uf.pk)
        self.assertTrue(res["ok"])
        self.assertEqual(res["text"], "plain text body")
        self.assertEqual(res["source"], "raw_file")

    def test_truncation_reports_full_length(self):
        uf = self._file(extracted_text="x" * 500)
        res = read_file(file_id=uf.pk, max_chars=100)
        self.assertTrue(res["ok"])
        self.assertEqual(len(res["text"]), 100)
        self.assertEqual(res["full_length"], 500)
        self.assertTrue(res["truncated"])

    def test_handler_never_raises(self):
        uf = self._file(name="report.docx", content=b"not really a docx")
        with patch("uploads.services._extract_text", side_effect=RuntimeError("boom")):
            res = read_file(file_id=uf.pk)
        self.assertFalse(res["ok"])
        self.assertIn("boom", res["error"])

    # ── non-PDF pagination (previously a silent no-op) ────────────────────────

    def test_non_pdf_page_from_returns_different_chunks(self):
        """
        page_from used to be silently ignored for anything but a PDF — a
        large CSV/XLSX/DOCX always returned the exact same first max_chars
        slice no matter what page was requested, so a document larger than
        max_chars was permanently stuck at its first chunk from the model's
        perspective even though the full text was sitting in the DB.
        """
        text = "".join(f"line{i}\n" for i in range(5000))
        uf = self._file(name="big.csv", extracted_text=text)

        page1 = read_file(file_id=uf.pk, max_chars=100, page_from=1)
        page2 = read_file(file_id=uf.pk, max_chars=100, page_from=2)

        self.assertTrue(page1["ok"])
        self.assertTrue(page2["ok"])
        self.assertNotEqual(page1["text"], page2["text"],
                             "page_from had no effect — same slice returned twice")
        self.assertEqual(page1["text"], text[:100])
        self.assertEqual(page2["text"], text[100:200])
        self.assertEqual(page1["page_from"], 1)
        self.assertEqual(page2["page_from"], 2)
        self.assertGreater(page1["page_count"], 1)
        self.assertEqual(page1["page_count"], page2["page_count"])

    def test_non_pdf_pagination_stops_at_the_last_page(self):
        uf = self._file(name="small.csv", extracted_text="x" * 50)
        res = read_file(file_id=uf.pk, max_chars=100, page_from=99)
        self.assertTrue(res["ok"])
        self.assertEqual(res["page_from"], 1, "out-of-range page_from should clamp to the last real page")
        self.assertEqual(res["page_count"], 1)
        self.assertEqual(res["text"], "x" * 50)

    def test_non_pdf_pagination_prefers_cached_extracted_text_over_reparsing(self):
        """Paging shouldn't re-parse the file from disk on every call when
        extracted_text is already cached — only the first, extraction-time
        parse should ever touch _extract_text for a given file."""
        uf = self._file(name="big.docx", extracted_text="x" * 1000)
        with patch("uploads.services._extract_text") as mock_extract:
            res = read_file(file_id=uf.pk, max_chars=100, page_from=3)
        mock_extract.assert_not_called()
        self.assertTrue(res["ok"])
        self.assertEqual(res["text"], ("x" * 1000)[200:300])

    def test_non_pdf_pagination_falls_back_to_live_read_when_uncached(self):
        """If extraction hasn't produced extracted_text yet, paging should
        still work by reading (and, for non-txt/csv, extracting) the raw
        file live, same as the unpaged fallback already does."""
        uf = self._file(name="big.docx", extracted_text="")
        with patch("uploads.services._extract_text", return_value="y" * 1000) as mock_extract:
            res = read_file(file_id=uf.pk, max_chars=100, page_from=2)
        mock_extract.assert_called_once()
        # max_chars=None means "give me everything" for the pagination path
        # to slice locally, not a second truncated read.
        self.assertIsNone(mock_extract.call_args.kwargs.get("max_chars"))
        self.assertTrue(res["ok"])
        self.assertEqual(res["text"], ("y" * 1000)[100:200])

    def test_default_max_chars_is_the_raised_settings_value_not_12000(self):
        """read_file's default per-call budget was raised from 12 000 to
        settings.READ_FILE_DEFAULT_MAX_CHARS (40 000) — fewer page_from
        round-trips needed to fully read a large document."""
        from django.conf import settings

        uf = self._file(extracted_text="x" * 50000)
        res = read_file(file_id=uf.pk)  # no max_chars passed — uses the default

        expected_default = getattr(settings, "READ_FILE_DEFAULT_MAX_CHARS", 40000)
        self.assertEqual(expected_default, 40000)
        self.assertEqual(len(res["text"]), expected_default)
        self.assertNotEqual(len(res["text"]), 12000)


class ExportFileTests(TestCase):
    """
    export_file(file_id, target_format) — convert an already-uploaded file's
    full extracted text into another format by id, without the LLM having to
    re-type the data through its own context (unlike write_xlsx). This is
    what makes "give me that file as a CSV/Excel file" possible for a file
    from an earlier turn the model has only ever seen a file_id for.
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="export@example.com", password="pw12345!",
            first_name="E", last_name="F",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    def _file(self, *, name="doc.csv", extracted_text="a,b,c\n1,2,3\n"):
        upload = SimpleUploadedFile(name, extracted_text.encode())
        return UploadedFile.objects.create(
            batch=self.batch, file=upload, original_filename=name,
            file_size_bytes=len(extracted_text), extension=name.rsplit(".", 1)[-1],
            parse_status="parsed", extracted_text=extracted_text,
        )

    def test_missing_file_is_reported(self):
        res = export_file(file_id=999999, target_format="csv")
        self.assertFalse(res["ok"])
        self.assertIn("not found", res["error"])

    def test_no_extracted_text_is_a_failure(self):
        uf = self._file(extracted_text="")
        res = export_file(file_id=uf.pk, target_format="csv")
        self.assertFalse(res["ok"])
        self.assertIn("no extracted text", res["error"].lower())

    def test_unsupported_format_is_rejected(self):
        uf = self._file()
        res = export_file(file_id=uf.pk, target_format="pdf")
        self.assertFalse(res["ok"])
        self.assertIn("Unsupported", res["error"])

    def test_export_is_byte_faithful_not_llm_paraphrased(self):
        """The whole point vs. write_xlsx: no LLM in the loop, so the output
        must be an exact, deterministic function of the stored text."""
        original = "id,amount\nINV-001,1000\nINV-002,2000\n"
        uf = self._file(name="invoices.csv", extracted_text=original)

        for target_format, expect_in in [
            ("csv", "INV-001"),
            ("json", "INV-002"),
            ("txt", "amount"),
        ]:
            with self.subTest(target_format=target_format):
                res = export_file(file_id=uf.pk, target_format=target_format)
                self.assertTrue(res["ok"], res.get("error"))
                self.assertEqual(res["format"], target_format)
                self.assertEqual(res["chars"], len(original))
                out_path = Path(res["output_filename"])
                self.assertTrue(out_path.exists())
                content = out_path.read_bytes()
                self.assertIn(expect_in.encode(), content)
                out_path.unlink()

    def test_export_writes_xlsx_with_correct_row_count(self):
        original = "id,amount\nINV-001,1000\nINV-002,2000\nINV-003,3000\n"
        uf = self._file(name="invoices.csv", extracted_text=original)
        res = export_file(file_id=uf.pk, target_format="xlsx")
        self.assertTrue(res["ok"], res.get("error"))
        out_path = Path(res["output_filename"])
        self.assertTrue(out_path.exists())

        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 4)  # header + 3 data rows
        self.assertEqual(rows[0], ("id", "amount"))
        # Values come from splitting plain text, not typed LLM output (unlike
        # write_xlsx) — always strings, which is correct here, not a bug.
        self.assertEqual(rows[1], ("INV-001", "1000"))
        out_path.unlink()

    def test_uses_original_filename_stem_not_file_id(self):
        uf = self._file(name="q3_report.csv", extracted_text="a,b\n1,2\n")
        res = export_file(file_id=uf.pk, target_format="json")
        self.assertTrue(res["ok"])
        self.assertTrue(Path(res["output_filename"]).name.startswith("q3_report"))
        Path(res["output_filename"]).unlink()


class DetectFileTypeTests(TestCase):
    """
    detect_file_type sniffed extracted_text and *persisted* the verdict. For a
    deferred large PDF that text is empty, so every big bill was permanently
    labelled "generic_pdf" and nothing ever recomputed it.
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="detect@example.com", password="pw12345!",
            first_name="D", last_name="T",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    def _pdf(self, **kwargs):
        upload = SimpleUploadedFile("bill.pdf", b"%PDF-1.4")
        defaults = dict(
            batch=self.batch, file=upload, original_filename="bill.pdf",
            file_size_bytes=8, extension="pdf", parse_status="pending",
            extracted_text="", extraction_deferred=True, page_count=300,
        )
        defaults.update(kwargs)
        return UploadedFile.objects.create(**defaults)

    def test_deferred_pdf_is_sniffed_live_not_mislabelled(self):
        uf = self._pdf()
        with patch("uploads.services.extract_pdf_page_range",
                   return_value={"ok": True, "text": "SAFARICOM POSTPAY TAX INVOICE"}):
            res = detect_file_type(file_id=uf.pk)

        self.assertEqual(res["detected_type"], "safaricom_bill")
        self.assertEqual(res["confidence"], "high")
        self.assertTrue(res["based_on_content"])
        uf.refresh_from_db()
        self.assertEqual(uf.detected_type, "safaricom_bill")

    def test_verdict_from_no_content_is_not_persisted(self):
        uf = self._pdf()
        with patch("uploads.services.extract_pdf_page_range",
                   return_value={"ok": True, "text": ""}):
            res = detect_file_type(file_id=uf.pk)

        self.assertFalse(res["based_on_content"])
        self.assertEqual(res["confidence"], "low")
        uf.refresh_from_db()
        self.assertEqual(uf.detected_type or "", "",
                         "provisional guess persisted as if it were real")

    def test_detection_is_redone_when_extraction_lands(self):
        uf = self._pdf(detected_type="generic_pdf", detection_confidence="low")
        UploadService.complete_extraction(
            uf.pk, text="SAFARICOM POSTPAY BILLED AMOUNT 1,234", page_count=300,
        )
        uf.refresh_from_db()
        self.assertEqual(uf.detected_type, "safaricom_bill",
                         "stale low-confidence verdict never recomputed")
        self.assertEqual(uf.detection_confidence, "high")

    def test_high_confidence_verdict_is_not_clobbered(self):
        uf = self._pdf(detected_type="acon_export", detection_confidence="high")
        UploadService.complete_extraction(uf.pk, text="SAFARICOM POSTPAY", page_count=1)
        uf.refresh_from_db()
        self.assertEqual(uf.detected_type, "acon_export")


class PromptTransformInputTests(TestCase):
    """
    _call_prompt_transform silently capped input at 10 pages / 8 000 chars and
    still returned ok=True, so a reconciliation could be built from ~3% of a
    document with nothing recording that.
    """

    def setUp(self):
        self.config = SimpleNamespace(
            name="extract_safaricom_bill",
            system_prompt="Extract from:\n{file_text}",
            output_schema=None,
        )
        self.user = get_user_model().objects.create_user(
            email="pt@example.com", password="pw12345!", first_name="P", last_name="T",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    def _pdf(self, **kwargs):
        upload = SimpleUploadedFile("bill.pdf", b"%PDF-1.4")
        defaults = dict(
            batch=self.batch, file=upload, original_filename="bill.pdf",
            file_size_bytes=8, extension="pdf", parse_status="pending",
            extracted_text="", page_count=300,
        )
        defaults.update(kwargs)
        return UploadedFile.objects.create(**defaults)

    def _run(self, arguments):
        """Call the dispatcher with a stubbed Grok client; return (result, system_prompt)."""
        captured = {}

        def _create(**kwargs):
            captured["system"] = kwargs["messages"][0]["content"]
            return SimpleNamespace(
                choices=[SimpleNamespace(message=SimpleNamespace(content="done"))],
                usage=SimpleNamespace(prompt_tokens=1, completion_tokens=1),
            )

        client = MagicMock()
        client.chat.completions.create.side_effect = _create
        with patch("tools.services._get_grok_client", return_value=client):
            res = _call_prompt_transform(self.config, arguments)
        return res, captured.get("system", "")

    def test_whole_document_is_read_not_the_first_ten_pages(self):
        uf = self._pdf()
        with patch("uploads.services.extract_pdf_page_range") as ranged:
            ranged.return_value = {"ok": True, "text": "body",
                                   "page_from": 1, "page_to": 300, "page_count": 300}
            res, _ = self._run({"file_id": uf.pk})

        self.assertIsNone(ranged.call_args.kwargs["page_to"],
                          "still capping the read at a fixed page count")
        self.assertTrue(res["ok"])
        self.assertEqual(res["page_count"], 300)
        self.assertFalse(res["input_truncated"])

    def test_explicit_page_range_is_honoured(self):
        uf = self._pdf()
        with patch("uploads.services.extract_pdf_page_range") as ranged:
            ranged.return_value = {"ok": True, "text": "body",
                                   "page_from": 40, "page_to": 60, "page_count": 300}
            self._run({"file_id": uf.pk, "page_from": 40, "page_to": 60})

        kwargs = ranged.call_args.kwargs
        self.assertEqual((kwargs["page_from"], kwargs["page_to"]), (40, 60))

    @override_settings(PROMPT_TRANSFORM_MAX_CHARS=100)
    def test_oversized_single_source_is_chunked_and_merged_not_truncated(self):
        """
        A single-source input over the cap is no longer silently truncated —
        it's split into <=max_chars pieces, each analyzed, then merged in
        pure Python (no further LLM call) into one result covering the whole
        document (see _run_chunked_prompt_transform). This config has no
        output_schema, so the merge is a labelled concatenation of every
        chunk's raw text rather than a structured JSON merge.
        """
        uf = self._pdf(extracted_text="x" * 5000)
        res, _ = self._run({"file_id": uf.pk})

        self.assertTrue(res["ok"])
        self.assertFalse(res["input_truncated"], "chunk+merge covers the whole document — nothing was dropped")
        self.assertEqual(res["input_full_length"], 5000)
        self.assertTrue(res.get("chunked"))
        self.assertEqual(res["chunk_count"], 50)  # 5000 chars / 100-char cap
        self.assertEqual(res["result"].count("done"), 50, "every chunk's response must survive the merge")
        self.assertIsNone(res["structured"], "no output_schema was set, so there is nothing to merge structurally")

    def test_untruncated_input_is_not_flagged(self):
        uf = self._pdf(extracted_text="short body")
        res, system_prompt = self._run({"file_id": uf.pk})
        self.assertFalse(res["input_truncated"])
        self.assertNotIn("TRUNCATED", system_prompt)
        self.assertIn("short body", system_prompt)


class ChunkedMergeTests(TestCase):
    """
    Regression tests for the map-reduce chunker's Python-merge fix. The old
    reduce step re-asked the LLM to re-emit the ENTIRE merged result in one
    4096-token completion — this is what silently collapsed large datasets
    (e.g. 1300 rows) down to a handful of returned rows even when the input
    side was fully covered by chunking. The fix merges each chunk's already-
    parsed structured JSON in Python instead, with no LLM call for the merge.
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="chunk@example.com", password="pw12345!", first_name="C", last_name="K",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    def _file(self, text, name="big.csv"):
        return UploadedFile.objects.create(
            batch=self.batch, file=SimpleUploadedFile(name, b"x"),
            original_filename=name, file_size_bytes=len(text),
            extension=name.rsplit(".", 1)[-1], parse_status="parsed",
            extracted_text=text,
        )

    EXTRACT_SCHEMA = {
        "type": "object",
        "properties": {
            "record_count": {"type": "integer"},
            "records": {"type": "array"},
        },
    }

    def _per_chunk_records_response(self, call_count: dict):
        """Each mocked completion returns 2 distinct records, keyed off the
        'part N of M' text _map_chunk always appends to the chunk prompt."""
        def _create(**kwargs):
            call_count["n"] += 1
            system = kwargs["messages"][0]["content"]
            m = re.search(r"part (\d+) of (\d+)", system)
            i = int(m.group(1))
            body = json.dumps({
                "record_count": 2,
                "records": [{"id": (i - 1) * 2}, {"id": (i - 1) * 2 + 1}],
            })
            return SimpleNamespace(
                choices=[SimpleNamespace(message=SimpleNamespace(content=body))],
                usage=SimpleNamespace(prompt_tokens=1, completion_tokens=1),
            )
        return _create

    @override_settings(PROMPT_TRANSFORM_MAX_CHARS=50)
    def test_chunked_prompt_transform_returns_all_rows_not_capped_subset(self):
        uf = self._file("x" * 500)  # 500 chars / 50-char cap -> 10 chunks
        config = SimpleNamespace(
            name="extract_invoice_data",
            system_prompt="Extract from:\n{file_text}",
            output_schema=self.EXTRACT_SCHEMA,
        )
        call_count = {"n": 0}
        client = MagicMock()
        client.chat.completions.create.side_effect = self._per_chunk_records_response(call_count)

        with patch("tools.services._get_grok_client", return_value=client):
            res = _call_prompt_transform(config, {"file_id": uf.pk})

        self.assertTrue(res["ok"])
        self.assertTrue(res["chunked"])
        chunk_count = res["chunk_count"]
        self.assertGreater(chunk_count, 1)
        self.assertEqual(len(res["structured"]["records"]), chunk_count * 2,
                          "every chunk's records must survive the merge, not just the last one")
        self.assertEqual(res["structured"]["record_count"], chunk_count * 2,
                          "the count must be recomputed from the merged array, never drift")

    @override_settings(PROMPT_TRANSFORM_MAX_CHARS=50)
    def test_reduce_step_makes_no_extra_llm_call(self):
        uf = self._file("x" * 500)
        config = SimpleNamespace(
            name="extract_invoice_data",
            system_prompt="Extract from:\n{file_text}",
            output_schema=self.EXTRACT_SCHEMA,
        )
        call_count = {"n": 0}
        client = MagicMock()
        client.chat.completions.create.side_effect = self._per_chunk_records_response(call_count)

        with patch("tools.services._get_grok_client", return_value=client):
            res = _call_prompt_transform(config, {"file_id": uf.pk})

        self.assertEqual(call_count["n"], res["chunk_count"],
                          "one completion per chunk, and no extra reduce completion")

    @override_settings(PROMPT_TRANSFORM_MAX_CHARS=80)
    def test_two_source_reconcile_chunking_covers_all_rows(self):
        big = self._file("A" * 800, name="big_a.txt")
        small = self._file("B" * 30, name="small_b.txt")
        config = SimpleNamespace(
            name="reconcile_datasets",
            system_prompt="A:\n{file_a_text}\nB:\n{file_b_text}",
            output_schema={
                "type": "object",
                "properties": {
                    "count_a": {"type": "integer"},
                    "count_b": {"type": "integer"},
                    "rows": {"type": "array"},
                },
            },
        )
        call_count = {"n": 0}

        def _create(**kwargs):
            call_count["n"] += 1
            system = kwargs["messages"][0]["content"]
            m = re.search(r"part (\d+) of (\d+)", system)
            i = int(m.group(1))
            body = json.dumps({
                "count_a": 1, "count_b": 999,  # every chunk re-reads the whole small side
                "rows": [{"id": f"row-{i}"}],
            })
            return SimpleNamespace(
                choices=[SimpleNamespace(message=SimpleNamespace(content=body))],
                usage=SimpleNamespace(prompt_tokens=1, completion_tokens=1),
            )

        client = MagicMock()
        client.chat.completions.create.side_effect = _create
        with patch("tools.services._get_grok_client", return_value=client):
            res = _call_prompt_transform(config, {"file_id_a": big.pk, "file_id_b": small.pk})

        self.assertTrue(res["ok"])
        self.assertTrue(res["chunked"])
        chunk_count = res["chunk_count"]
        self.assertGreater(chunk_count, 1)
        self.assertEqual(len(res["structured"]["rows"]), chunk_count,
                          "one row per chunk from the chunked (larger) side must all survive")
        self.assertEqual(res["structured"]["count_b"], 999,
                          "the unchunked side's own count must not be summed once per chunk")

    def test_single_shot_uses_raised_token_cap(self):
        uf = self._file("short text")
        config = SimpleNamespace(
            name="extract_invoice_data",
            system_prompt="Extract from:\n{file_text}",
            output_schema=self.EXTRACT_SCHEMA,
        )
        captured = {}

        def _create(**kwargs):
            captured["max_tokens"] = kwargs["max_tokens"]
            return SimpleNamespace(
                choices=[SimpleNamespace(message=SimpleNamespace(content='{"record_count":0,"records":[]}'))],
                usage=SimpleNamespace(prompt_tokens=1, completion_tokens=1),
            )

        client = MagicMock()
        client.chat.completions.create.side_effect = _create
        with patch("tools.services._get_grok_client", return_value=client):
            _call_prompt_transform(config, {"file_id": uf.pk})

        from django.conf import settings
        self.assertEqual(captured["max_tokens"], getattr(settings, "PROMPT_TRANSFORM_MAX_TOKENS", 12000))
        self.assertNotEqual(captured["max_tokens"], 4096)


class MultiFileResolutionTests(TestCase):
    """
    Only `file_id` was ever resolved, so every multi-file tool received an empty
    {file_text} and answered from nothing: reconcile_datasets (file_id_a /
    file_id_b) and summarise_batch (batch_id) were both non-functional.
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="multi@example.com", password="pw12345!", first_name="M", last_name="F",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    def _file(self, text, name="f.txt"):
        return UploadedFile.objects.create(
            batch=self.batch, file=SimpleUploadedFile(name, b"x"),
            original_filename=name, file_size_bytes=1,
            extension=name.rsplit(".", 1)[-1], parse_status="parsed",
            extracted_text=text,
        )

    def _run(self, config, arguments):
        captured = {}

        def _create(**kwargs):
            captured["system"] = kwargs["messages"][0]["content"]
            return SimpleNamespace(
                choices=[SimpleNamespace(message=SimpleNamespace(content="done"))],
                usage=SimpleNamespace(prompt_tokens=1, completion_tokens=1),
            )

        client = MagicMock()
        client.chat.completions.create.side_effect = _create
        with patch("tools.services._get_grok_client", return_value=client):
            res = _call_prompt_transform(config, arguments)
        return res, captured.get("system", "")

    def test_both_sides_of_a_reconciliation_are_loaded(self):
        a = self._file("URA SIDE CU-111", "ura.txt")
        b = self._file("ACON SIDE CU-111", "acon.txt")
        cfg = SimpleNamespace(
            name="reconcile_datasets", output_schema=None,
            system_prompt="A:\n{file_a_text}\nB:\n{file_b_text}",
        )
        res, prompt = self._run(cfg, {"file_id_a": a.pk, "file_id_b": b.pk})

        self.assertIn("URA SIDE CU-111", prompt)
        self.assertIn("ACON SIDE CU-111", prompt, "file B never reached the model")
        self.assertEqual(res["input_full_length"], len("URA SIDE CU-111") + len("ACON SIDE CU-111"))

    def test_batch_id_loads_every_file_labelled(self):
        self._file("first doc", "a.txt")
        self._file("second doc", "b.txt")
        cfg = SimpleNamespace(
            name="summarise_batch", output_schema=None,
            system_prompt="Batch:\n{batch_text}",
        )
        _, prompt = self._run(cfg, {"batch_id": self.batch.pk})

        self.assertIn("first doc", prompt)
        self.assertIn("second doc", prompt)
        self.assertIn("a.txt", prompt, "files are not labelled in the batch blob")

    def test_unreferenced_placeholders_do_not_leak_into_the_prompt(self):
        a = self._file("only A", "a.txt")
        cfg = SimpleNamespace(
            name="reconcile_datasets", output_schema=None,
            system_prompt="A:\n{file_a_text}\nB:\n{file_b_text}",
        )
        _, prompt = self._run(cfg, {"file_id_a": a.pk})
        self.assertNotIn("{file_b_text}", prompt,
                         "literal placeholder reached the model as content")

    @override_settings(PROMPT_TRANSFORM_MAX_CHARS=100)
    def test_oversized_two_source_input_is_chunked_not_truncated(self):
        """
        Two equally-sized oversized sources used to be silently truncated to
        their per-source budget share. Now the larger side (ties go to the
        first-declared source) is chunked and merged instead, so the whole
        pair is covered rather than a fixed-budget slice of each.
        """
        a = self._file("a" * 500, "a.txt")
        b = self._file("b" * 500, "b.txt")
        cfg = SimpleNamespace(
            name="reconcile_datasets", output_schema=None,
            system_prompt="A:\n{file_a_text}\nB:\n{file_b_text}",
        )
        res, _ = self._run(cfg, {"file_id_a": a.pk, "file_id_b": b.pk})

        self.assertTrue(res["ok"])
        self.assertFalse(res["input_truncated"], "chunk+merge now covers both sources — nothing was dropped")
        self.assertEqual(res["input_full_length"], 1000)
        self.assertTrue(res.get("chunked"))
        self.assertGreater(res["chunk_count"], 1)

    def test_single_file_contract_still_works(self):
        f = self._file("legacy body", "a.txt")
        cfg = SimpleNamespace(
            name="extract_invoice_data", output_schema=None,
            system_prompt="Doc:\n{file_text}",
        )
        _, prompt = self._run(cfg, {"file_id": f.pk})
        self.assertIn("legacy body", prompt)

    def test_file_text_falls_back_to_side_a_for_legacy_prompts(self):
        a = self._file("side A body", "a.txt")
        b = self._file("side B body", "b.txt")
        cfg = SimpleNamespace(
            name="reconcile_datasets", output_schema=None,
            system_prompt="Doc:\n{file_text}",
        )
        _, prompt = self._run(cfg, {"file_id_a": a.pk, "file_id_b": b.pk})
        self.assertIn("side A body", prompt)


class RestoredDomainHandlerTests(TestCase):
    """
    The CU-matching logic these cover is exactly what the generic
    prompt_transform replacements could not do:
      - URA/ACON join on the fiscal/statutory column, NOT ACON's Item Number
      - identifier normalisation so Excel's '12345.0' matches '12345'
      - URA number format '4 862 563,00' (space thousands, comma decimal)
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="domain@example.com", password="pw12345!", first_name="D", last_name="H",
        )
        self.batch = UploadService.create_batch(label="t", user=self.user)

    # ── normalisation primitives ──────────────────────────────────────────────

    def test_norm_id_strips_excel_float_artefact(self):
        from tools.handlers import _norm_id
        self.assertEqual(_norm_id("0012345678.0"), "0012345678")
        self.assertEqual(_norm_id(" 12345 "), "12345")
        self.assertEqual(_norm_id(12345), "12345")

    def test_norm_number_handles_ura_and_standard_formats(self):
        from tools.handlers import _norm_number
        self.assertEqual(_norm_number("4 862 563,00"), 4862563.0)   # URA style
        self.assertEqual(_norm_number("1,234.56"), 1234.56)         # standard
        self.assertEqual(_norm_number("1000"), 1000.0)

    def test_saf_department_strips_company_prefix(self):
        from tools.handlers import _saf_department
        self.assertEqual(_saf_department("Kuehne + Nagel Ltd - Finance"), "Finance")
        self.assertEqual(_saf_department("Kuehne and Nagel Airfreight"), "Airfreight")
        self.assertEqual(_saf_department("Standalone Name"), "Standalone Name")

    # ── the URA/ACON join ─────────────────────────────────────────────────────

    def _xlsx(self, name, headers, rows):
        import openpyxl, io
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return UploadedFile.objects.create(
            batch=self.batch, file=SimpleUploadedFile(name, buf.read()),
            original_filename=name, file_size_bytes=1, extension="xlsx",
            parse_status="parsed",
        )

    def test_reconciliation_joins_on_statutory_column_not_item_number(self):
        from tools.handlers import reconcile_ura_vs_acon

        ura = self._xlsx(
            "ura.xlsx",
            ["FDN", "Name of Purchaser", "Invoice Date", "TOTAL (A+B)", "VAT Charged"],
            [["0012345678", "ACME LTD", "01-08-2026", 1000.0, 100.0]],
        )
        # Item Number is a decoy: it must NOT be used as the join key.
        acon = self._xlsx(
            "acon.xlsx",
            ["Debtor Account", "Full Name", "Item Number",
             "Statutory Item No(For Download VAT)", "LC Amount"],
            [["D1", "ACME LTD", "999999", "0012345678", 1000.0]],
        )

        res = reconcile_ura_vs_acon(ura_file_id=ura.pk, acon_file_id=acon.pk)
        self.assertTrue(res["ok"], res.get("error"))
        self.assertEqual(res["matched_count"], 1, "joined on the wrong column")
        self.assertEqual(res["missing_in_acon"], 0)
        self.assertEqual(res["variance_count"], 0)
        self.assertEqual(res["rows"][0]["acon_item"], "999999")

    def test_excel_float_artefact_still_matches(self):
        from tools.handlers import reconcile_ura_vs_acon
        ura  = self._xlsx("ura.xlsx",
                          ["FDN", "Name of Purchaser", "Invoice Date", "TOTAL (A+B)"],
                          [["12345", "ACME", "01-08-2026", 500.0]])
        acon = self._xlsx("acon.xlsx",
                          ["Statutory Item No(For Download VAT)", "Full Name", "LC Amount"],
                          [[12345.0, "ACME", 500.0]])   # Excel numeric cell -> '12345.0'

        res = reconcile_ura_vs_acon(ura_file_id=ura.pk, acon_file_id=acon.pk)
        self.assertEqual(res["matched_count"], 1,
                         "'12345.0' vs '12345' failed to match — _norm_id not applied")

    def test_variance_and_missing_rows_are_classified(self):
        from tools.handlers import reconcile_ura_vs_acon
        ura = self._xlsx("ura.xlsx",
                         ["FDN", "Name of Purchaser", "Invoice Date", "TOTAL (A+B)"],
                         [["111", "A", "01-08-2026", 1000.0],
                          ["222", "B", "01-08-2026", 2000.0],
                          ["333", "C", "01-08-2026", 3000.0]])
        acon = self._xlsx("acon.xlsx",
                          ["Statutory Item No(For Download VAT)", "Full Name", "LC Amount"],
                          [["111", "A", 1000.0],      # match
                           ["222", "B", 1900.0],      # variance
                           ["444", "D", 4000.0]])     # missing in URA

        res = reconcile_ura_vs_acon(ura_file_id=ura.pk, acon_file_id=acon.pk)
        self.assertEqual(res["matched_count"], 2)
        self.assertEqual(res["variance_count"], 1)
        self.assertEqual(res["missing_in_acon"], 1)   # 333
        self.assertEqual(res["missing_in_ura"], 1)    # 444

    def test_workpaper_is_written_and_collectable(self):
        from pathlib import Path
        from tools.handlers import reconcile_ura_vs_acon
        ura  = self._xlsx("ura.xlsx",
                          ["FDN", "Name of Purchaser", "Invoice Date", "TOTAL (A+B)"],
                          [["111", "A", "01-08-2026", 10.0]])
        acon = self._xlsx("acon.xlsx",
                          ["Statutory Item No(For Download VAT)", "Full Name", "LC Amount"],
                          [["111", "A", 10.0]])
        res = reconcile_ura_vs_acon(ura_file_id=ura.pk, acon_file_id=acon.pk)
        # collect_output_files reads result["output_filename"] as an absolute path.
        self.assertTrue(Path(res["output_filename"]).exists())

    def test_txt_fiscal_side_parses_cu_blocks(self):
        from tools.handlers import _parse_fiscal_side
        raw = (
            "FISCAL RECEIPT\r\n"
            "CU INVOICE NUMBER: 0012345678\r\n"
            "01-08-2026 10:30:00\r\n"
            "TOTAL:   4 862 563,00\r\n"
            "TAXES:   662 563,00\r\n"
            "CREDIT NOTE\r\n"
            "CU INVOICE NUMBER: 0087654321\r\n"
            "02-08-2026 11:00:00\r\n"
            "TOTAL:   1 000,00\r\n"
            "TAXES:   100,00\r\n"
        )
        uf = UploadedFile.objects.create(
            batch=self.batch, file=SimpleUploadedFile("ura.txt", raw.encode()),
            original_filename="ura.txt", file_size_bytes=len(raw), extension="txt",
            parse_status="parsed",
        )
        recs = _parse_fiscal_side(uf)
        self.assertEqual(len(recs), 2, "credit notes or CU blocks not parsed")
        self.assertEqual(recs[0]["fiscal_no"], "0012345678")
        self.assertEqual(recs[0]["total"], 4862563.0)


class OrgIsolationTests(TestCase):
    """
    tools/views.py had the most severe pre-existing leaks in the whole
    audit: ToolCallListView/ToolCallDetailView exposed every user's actual
    extracted-data payloads (arguments/result) platform-wide with zero
    scoping, and ToolDefinitionListView listed every user's custom tools
    together. Phase 3 fixes both by org-scoping, while also making custom
    tools genuinely org-SHARED (any org member can list/test; only the
    creator or an org-admin can edit/delete).
    """

    def setUp(self):
        User = get_user_model()
        self.org_a = Organization.objects.create(name="Org A")
        self.org_b = Organization.objects.create(name="Org B")
        self.creator_a = User.objects.create_user(
            email="creator_a@example.com", password="pw12345!",
            first_name="C", last_name="A", organization=self.org_a,
        )
        self.teammate_a = User.objects.create_user(
            email="teammate_a@example.com", password="pw12345!",
            first_name="T", last_name="A", organization=self.org_a,
        )
        self.admin_a = User.objects.create_user(
            email="admin_a@example.com", password="pw12345!",
            first_name="Ad", last_name="A", role="admin", organization=self.org_a,
        )
        self.user_b = User.objects.create_user(
            email="user_b@example.com", password="pw12345!",
            first_name="U", last_name="B", organization=self.org_b,
        )

        self.custom_tool = ToolDefinition.objects.create(
            name="org_a_webhook_tool", display_name="Org A Webhook Tool",
            description="test", category="utility", tool_type="webhook",
            handler="", enabled=True, is_safe=True,
            created_by=self.creator_a, organization=self.org_a,
        )
        from tools.models import UserToolConfig
        UserToolConfig.objects.create(tool=self.custom_tool, webhook_url="https://example.com/hook")

        self.call = ToolCall.objects.create(
            tool=self.custom_tool, arguments={"x": 1}, result={"ok": True, "secret": "org-a-data"},
            status="success", organization=self.org_a,
        )

        self.client = APIClient()

    def test_tool_call_list_is_org_scoped(self):
        """The most severe leak found in the audit — real extracted data
        must never be visible cross-org."""
        self.client.force_authenticate(user=self.teammate_a)
        resp = self.client.get("/api/tools/calls/")
        self.assertEqual(resp.status_code, 200)
        ids = [c["id"] for c in resp.json()]
        self.assertIn(self.call.pk, ids)

        self.client.force_authenticate(user=self.user_b)
        resp = self.client.get("/api/tools/calls/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), [])

    def test_tool_call_detail_404s_cross_org_instead_of_leaking(self):
        self.client.force_authenticate(user=self.user_b)
        resp = self.client.get(f"/api/tools/calls/{self.call.pk}/")
        self.assertEqual(resp.status_code, 404)

    def test_custom_tool_is_listed_for_org_members_not_other_orgs(self):
        self.client.force_authenticate(user=self.teammate_a)
        resp = self.client.get("/api/tools/custom/")
        self.assertEqual(resp.status_code, 200)
        names = [t["name"] for t in resp.json()]
        self.assertIn("org_a_webhook_tool", names)

        self.client.force_authenticate(user=self.user_b)
        resp = self.client.get("/api/tools/custom/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), [])

    def test_non_creator_org_member_can_test_but_not_edit(self):
        self.client.force_authenticate(user=self.teammate_a)

        with patch("tools.services._call_webhook", return_value={"ok": True}):
            test_resp = self.client.post(
                f"/api/tools/custom/{self.custom_tool.pk}/test/", {"arguments": {}}, format="json",
            )
        self.assertEqual(test_resp.status_code, 200)

        patch_resp = self.client.patch(
            f"/api/tools/custom/{self.custom_tool.pk}/", {"description": "hijacked"}, format="json",
        )
        self.assertEqual(patch_resp.status_code, 404)

    def test_org_admin_can_edit_even_if_not_creator(self):
        self.client.force_authenticate(user=self.admin_a)
        patch_resp = self.client.patch(
            f"/api/tools/custom/{self.custom_tool.pk}/", {"description": "updated by admin"}, format="json",
        )
        self.assertEqual(patch_resp.status_code, 200)

    def test_cross_org_user_cannot_test_or_view(self):
        self.client.force_authenticate(user=self.user_b)
        resp = self.client.get(f"/api/tools/custom/{self.custom_tool.pk}/")
        self.assertEqual(resp.status_code, 404)
        resp = self.client.post(
            f"/api/tools/custom/{self.custom_tool.pk}/test/", {"arguments": {}}, format="json",
        )
        self.assertEqual(resp.status_code, 404)

    def test_llm_tool_discovery_excludes_other_orgs_custom_tools(self):
        """The agent's own tool-selection (_load_tool_schemas/_load_tool_map)
        must not offer, or be able to invoke, another org's custom tool —
        found while implementing Phase 3, not originally in the audit."""
        from tools.services import _load_tool_map, _load_tool_schemas

        names = {s["function"]["name"] for s in _load_tool_schemas(organization=self.org_b)}
        self.assertNotIn("org_a_webhook_tool", names)

        tool_map = _load_tool_map(organization=self.org_b)
        self.assertNotIn("org_a_webhook_tool", tool_map)

        # Same org: visible.
        names_a = {s["function"]["name"] for s in _load_tool_schemas(organization=self.org_a)}
        self.assertIn("org_a_webhook_tool", names_a)


class ToolCallOutputDownloadTests(TestCase):
    """
    ToolRunView/CustomToolTestView only ever returned {tool_call_id, result}
    with a raw server filesystem path in result["output_filename"] — no
    endpoint anywhere let a non-chat caller actually download it. This is
    the new download surface that closes that gap.
    """

    def setUp(self):
        self.org_a = Organization.objects.create(name="Org A")
        self.org_b = Organization.objects.create(name="Org B")
        self.user_a = get_user_model().objects.create_user(
            email="user_a@example.com", password="pw12345!",
            first_name="U", last_name="A", organization=self.org_a,
        )
        self.user_b = get_user_model().objects.create_user(
            email="user_b@example.com", password="pw12345!",
            first_name="U", last_name="B", organization=self.org_b,
        )
        self.tool = ToolDefinition.objects.create(
            name="write_xlsx", display_name="Write Xlsx", description="test",
            category="report", tool_type="builtin", handler="", enabled=True, is_safe=True,
        )
        self.client = APIClient()

    def _write_temp_output(self, tmp_path, content=b"fake xlsx bytes"):
        tmp_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path.write_bytes(content)

    def test_same_org_can_download_output(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp_dir:
            out_path = Path(tmp_dir) / "report.xlsx"
            self._write_temp_output(out_path, b"hello xlsx")
            call = ToolCall.objects.create(
                tool=self.tool, arguments={}, result={"ok": True, "output_filename": str(out_path)},
                status="success", organization=self.org_a,
            )
            self.client.force_authenticate(user=self.user_a)
            resp = self.client.get(f"/api/tools/calls/{call.pk}/output/download/")
            self.assertEqual(resp.status_code, 200)
            content = b"".join(resp.streaming_content)
            self.assertEqual(content, b"hello xlsx")

    def test_cross_org_download_404s(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp_dir:
            out_path = Path(tmp_dir) / "report.xlsx"
            self._write_temp_output(out_path)
            call = ToolCall.objects.create(
                tool=self.tool, arguments={}, result={"ok": True, "output_filename": str(out_path)},
                status="success", organization=self.org_a,
            )
            self.client.force_authenticate(user=self.user_b)
            resp = self.client.get(f"/api/tools/calls/{call.pk}/output/download/")
            self.assertEqual(resp.status_code, 404)

    def test_download_404s_when_no_output_filename(self):
        call = ToolCall.objects.create(
            tool=self.tool, arguments={}, result={"ok": True},
            status="success", organization=self.org_a,
        )
        self.client.force_authenticate(user=self.user_a)
        resp = self.client.get(f"/api/tools/calls/{call.pk}/output/download/")
        self.assertEqual(resp.status_code, 404)

    def test_download_404s_not_500s_when_file_missing_on_disk(self):
        call = ToolCall.objects.create(
            tool=self.tool, arguments={},
            result={"ok": True, "output_filename": "/nonexistent/path/report.xlsx"},
            status="success", organization=self.org_a,
        )
        self.client.force_authenticate(user=self.user_a)
        resp = self.client.get(f"/api/tools/calls/{call.pk}/output/download/")
        self.assertEqual(resp.status_code, 404)
