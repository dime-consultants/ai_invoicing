# ai_engine/unstructured_pipeline.py
import logging
import json
from io import BytesIO
from pathlib import Path

import openpyxl
import pdfplumber
from django.core.files.base import ContentFile

from .services import _get_client, GROK_MODEL

logger = logging.getLogger(__name__)


class UnstructuredPipeline:
    """
    Intelligent file → structured data pipeline using Grok.
    """

    @staticmethod
    def _read_file(file_obj, filename: str) -> dict:
        """Extract raw text + metadata from various file types."""
        ext = Path(filename).suffix.lower()
        content = ""
        metadata = {"filename": filename, "file_type": ext}

        try:
            if ext == '.txt':
                content = file_obj.read().decode('utf-8', errors='replace')

            elif ext == '.pdf':
                file_obj.seek(0)
                with pdfplumber.open(file_obj) as pdf:
                    for i, page in enumerate(pdf.pages):
                        text = page.extract_text() or ""
                        content += f"\n--- Page {i+1} ---\n{text}\n"
                        tables = page.extract_tables()
                        if tables:
                            content += f"\n[TABLE on Page {i+1}]\n"

            elif ext in ('.xlsx', '.xls'):
                file_obj.seek(0)
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    content += f"\n=== Sheet: {sheet} ===\n"
                    for row in ws.iter_rows(values_only=True):
                        content += str(row) + "\n"

            elif ext == '.csv':
                content = file_obj.read().decode('utf-8', errors='replace')

            else:
                content = f"[Unsupported format: {ext}]"

        except Exception as e:
            logger.exception(f"Error reading {filename}")
            content = f"[ERROR READING FILE: {str(e)}]"

        return {
            "content": content[:80000],  # Truncate for token limit
            "metadata": metadata,
            "raw_length": len(content)
        }

    @staticmethod
    def extract_structured_data(file_obj, filename: str, domain_context: str = "") -> dict:
        """
        Main entry point: File → Clean Structured JSON
        """
        try:
            file_data = UnstructuredPipeline._read_file(file_obj, filename)

            system_prompt = """
You are an expert financial data extraction specialist for East Africa (Uganda/Kenya).
Extract all structured information from the document with high accuracy.
Always return valid JSON. Never add explanations outside the JSON.
"""

            user_prompt = f"""
Domain Context: {domain_context or 'Financial / Tax / Invoice documents'}

File: {filename}

--- DOCUMENT CONTENT ---
{file_data['content']}
--- END DOCUMENT ---

Extract the data into this schema (add more fields if needed):

{{
  "document_type": "URA_FISCAL_RECEIPT | KN_PERIODICAL_REPORT | SAFARICOM_INVOICE | ACON_SALES | OTHER",
  "company": "...",
  "period_from": "YYYY-MM-DD",
  "period_to": "YYYY-MM-DD",
  "records": [
    {{
      "type": "FISCAL_RECEIPT | CREDIT_NOTE",
      "cu_invoice_number": "00400...",
      "date": "YYYY-MM-DD",
      "total_amount": 12345.67,
      "tax_amount": 1234.56,
      "net_amount": 11111.11,
      "currency": "UGX | KES"
    }}
  ],
  "summary": {{
    "total_records": 45,
    "total_amount": 987654.32,
    "total_tax": 123456.78
  }}
}}
"""

            client = _get_client()

            response = client.chat.completions.create(
                model=GROK_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.0,
                max_tokens=8192,
            )

            raw = response.choices[0].message.content
            structured = UnstructuredPipeline._safe_json_parse(raw)

            logger.info(f"Successfully extracted structured data from {filename}")
            return {
                "success": True,
                "filename": filename,
                "structured": structured,
                "raw_response": raw,
                "metadata": file_data["metadata"]
            }

        except Exception as e:
            logger.exception("Unstructured pipeline failed")
            return {
                "success": False,
                "error": str(e),
                "filename": filename
            }

    @staticmethod
    def _safe_json_parse(text: str):
        """Clean and parse LLM JSON output."""
        text = text.strip()
        if text.startswith("```"):
            text = text.split("```", 2)[-2].strip()
            if text.startswith("json"):
                text = text[4:].strip()
        try:
            return json.loads(text)
        except:
            return {"error": "JSON parsing failed", "raw": text[:1000]}