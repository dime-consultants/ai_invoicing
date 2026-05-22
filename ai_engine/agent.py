# ai_engine/agent.py
"""
InvoiceAgent
============
Orchestrates end-to-end invoice processing for chat-uploaded files.

Flow
----
1. Read every attached file into text (reuses _read_file_to_text from services).
2. Call Grok with AGENT_SYSTEM_PROMPT + tool manifest to get a structured plan.
3. Execute each step in the plan:
   a. Extraction  — local parser builds the primary xlsx from file content.
   b. Conversion  — ConversionService tools (txt_to_xlsx, pdf_to_xlsx, xlsx_clean).
   c. Report      — ReportService to generate a summary xlsx.
   d. Analysis    — AIEngineService for flag_anomalies, summarise_batch, etc.
4. Collect all output BytesIO objects and return them alongside the response text.

All DB work (batch / UploadBatch creation, record persistence) is handled by the
existing UploadService / ConversionService / ReconciliationService / ReportService
so this class only orchestrates — it never talks to the DB directly.
"""

from __future__ import annotations

import json
import logging
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from django.utils import timezone

from .prompts import AGENT_SYSTEM_PROMPT, WORKFLOW_OVERRIDES

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalise_number(raw) -> float | str:
    """
    Convert ambiguous number strings to float.
    Handles:
      • space-as-thousands + comma-decimal  →  "4 862 563,00"  →  4862563.0
      • standard comma-thousands + dot-decimal  →  "1,234.56"  →  1234.56
      • plain strings that are not numbers    →  returned as-is
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    # Space thousands + comma decimal  (European / URA style)
    if re.match(r"^[\d ]+,\d{1,2}$", s):
        s = s.replace(" ", "").replace(",", ".")
    else:
        # Standard: remove comma-thousands, keep dot-decimal
        s = s.replace(",", "").replace(" ", "")

    try:
        return float(s)
    except (ValueError, TypeError):
        return raw


def _strip_json_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.rsplit("```", 1)[0].strip()
    return text


def _build_xlsx(
    headers: list,
    rows: list,
    sheet_name: str = "Data",
    extra_sheets: list[dict] | None = None,
) -> BytesIO:
    """
    Build an xlsx workbook.

    extra_sheets: list of {"title": str, "headers": list, "rows": list}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_sheet(worksheet, hdrs, data_rows):
        for col_idx, header in enumerate(hdrs, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=str(header))
            cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align

        for row_idx, row in enumerate(data_rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if value is not None and str(value).strip():
                    normalised = _normalise_number(value)
                    cell.value = normalised
                else:
                    cell.value = "" if value is None else str(value)
                cell.alignment = Alignment(horizontal="left", vertical="top")

        # Auto-width (capped at 50)
        for col_idx, header in enumerate(hdrs, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = len(str(header))
            for row_idx in range(2, min(len(data_rows) + 2, 200)):
                val = worksheet.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(val or "")))
            worksheet.column_dimensions[col_letter].width = min(max_len + 3, 50)

    _write_sheet(ws, headers, rows)

    for extra in (extra_sheets or []):
        new_ws = wb.create_sheet(title=extra["title"][:31])
        _write_sheet(new_ws, extra["headers"], extra["rows"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Report builders (local — no DB required)
# ---------------------------------------------------------------------------

def _build_summary_sheet(
    plan: dict,
    extraction: dict,
    file_count: int,
) -> dict:
    """
    Build a 'Summary' sheet dict to attach to any report xlsx.
    Returns an extra_sheet dict ready for _build_xlsx.
    """
    rows = [
        ["Plan",          plan.get("plan_summary", "")],
        ["Files processed", file_count],
        ["Records extracted", extraction.get("record_count") or len(extraction.get("rows", []))],
        ["Output file",   extraction.get("filename", "")],
        ["Notes",         plan.get("notes", "")],
        ["Generated at",  timezone.now().isoformat()],
    ]
    return {
        "title": "Summary",
        "headers": ["Field", "Value"],
        "rows": rows,
    }


def _build_report_xlsx(
    plan: dict,
    extraction: dict,
    file_count: int,
) -> tuple[BytesIO, str]:
    """
    Build a standalone report xlsx that contains:
      Sheet 1 — Summary  (plan, metadata, timestamp)
      Sheet 2 — Extracted Data  (same headers/rows as primary output)
    """
    ext = extraction or {}
    headers = ext.get("headers") or ["No Data"]
    rows    = ext.get("rows")    or []
    primary_filename = ext.get("filename", "extracted_data.xlsx")
    base    = Path(primary_filename).stem

    summary_sheet = _build_summary_sheet(plan, ext, file_count)

    buf = _build_xlsx(
        headers=["Field", "Value"],
        rows=summary_sheet["rows"],
        sheet_name="Summary",
        extra_sheets=[
            {
                "title": "Extracted Data",
                "headers": headers,
                "rows": rows,
            }
        ],
    )

    report_filename = f"{base}_report.xlsx"
    return buf, report_filename


# ---------------------------------------------------------------------------
# InvoiceAgent
# ---------------------------------------------------------------------------

class InvoiceAgent:
    """
    Main agent class.  Call .run() to process one or more attachments.

    Returns
    -------
    (response_text: str, output_files: list[dict])

    Each output_file dict:
        {
            "filename":     str,
            "content":      BytesIO,
            "content_type": str,
        }
    """

    XLSX_CONTENT_TYPE = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    def __init__(
        self,
        user,
        message: str,
        file_attachments: list,
        workflow_id: int | None = None,
        workflow_option: str | None = None,
        workflow_type: str | None = None,
        conversation_history: list[dict] | None = None,
    ):
        self.user               = user
        self.message            = message
        self.file_attachments   = file_attachments or []
        self.workflow_id        = workflow_id
        self.workflow_option    = workflow_option
        self.workflow_type      = workflow_type  # e.g. "ura_processing"
        self.conversation_history = conversation_history or []

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def run(self) -> tuple[str, list[dict]]:
        if not self.file_attachments:
            return self._no_files_response(), []

        # 1. Read files into text for the AI
        file_texts = self._read_attachments()

        # 2. Ask the AI for a plan + extraction JSON
        raw_plan = self._call_ai(file_texts)

        # 3. Parse the plan
        plan = self._parse_plan(raw_plan)

        # 4. Execute the plan and collect output files
        output_files, response_text = self._execute_plan(plan)

        return response_text, output_files

    # ------------------------------------------------------------------
    # Step 1 — read attachments
    # ------------------------------------------------------------------

    def _read_attachments(self) -> list[str]:
        from .services import _read_file_to_text  # avoid circular at module level
        texts = []
        for att in self.file_attachments:
            try:
                content = _read_file_to_text(att.file.open("rb"), att.filename)
                texts.append(content)
            except Exception as exc:
                logger.warning("Could not read %s: %s", att.filename, exc)
                texts.append(f"[Could not read {att.filename}: {exc}]")
        return texts

    # ------------------------------------------------------------------
    # Step 2 — call AI
    # ------------------------------------------------------------------

    def _build_system_prompt(self) -> str:
        override = ""
        if self.workflow_type and self.workflow_type in WORKFLOW_OVERRIDES:
            override = WORKFLOW_OVERRIDES[self.workflow_type]
        elif self.workflow_option == "full_pipeline":
            override = (
                "Run the complete pipeline: convert → reconcile → report → AI analysis. "
                "Set both report_requested and analysis_requested to true.\n\n"
            )
        return override + AGENT_SYSTEM_PROMPT

    def _call_ai(self, file_texts: list[str]) -> str:
        from .services import generate_chat_response  # lazy import

        file_section = "\n\n---\n\n".join(file_texts)
        user_message = (
            f"{self.message}\n\n"
            "=== ATTACHED FILE CONTENT ===\n"
            f"{file_section}"
        )

        system_prompt = self._build_system_prompt()

        try:
            return generate_chat_response(
                user_message=user_message,
                system_prompt=system_prompt,
                conversation_history=self.conversation_history,
            )
        except Exception as exc:
            logger.exception("Agent AI call failed: %s", exc)
            return json.dumps({
                "plan_summary": "AI call failed — falling back to local extraction.",
                "steps": [{"step": 1, "tool": "local_fallback", "reason": str(exc), "file_hint": None}],
                "extraction": {},
                "report_requested": True,
                "analysis_requested": False,
                "notes": str(exc),
            })

    # ------------------------------------------------------------------
    # Step 3 — parse plan JSON
    # ------------------------------------------------------------------

    def _parse_plan(self, raw: str) -> dict:
        try:
            cleaned = _strip_json_fences(raw)
            return json.loads(cleaned)
        except json.JSONDecodeError:
            logger.warning("Agent: could not parse AI plan JSON; using empty plan")
            return {
                "plan_summary": "Could not parse AI plan.",
                "steps": [],
                "extraction": {},
                "report_requested": True,
                "analysis_requested": False,
                "notes": raw[:500],
            }

    # ------------------------------------------------------------------
    # Step 4 — execute plan
    # ------------------------------------------------------------------

    def _execute_plan(self, plan: dict) -> tuple[list[dict], str]:
        output_files: list[dict] = []
        messages: list[str] = []

        extraction = plan.get("extraction") or {}
        headers    = extraction.get("headers") or []
        rows       = extraction.get("rows")    or []
        filename   = extraction.get("filename") or "extracted_data.xlsx"
        summary    = extraction.get("summary") or plan.get("plan_summary", "")
        record_count = extraction.get("record_count") or len(rows)

        # ── 4a. Primary extraction xlsx ───────────────────────────────
        if headers and rows:
            # Normalise all numeric values in rows
            clean_rows = [
                [_normalise_number(cell) for cell in row]
                for row in rows
            ]
            if not filename.endswith(".xlsx"):
                filename = Path(filename).stem + ".xlsx"

            primary_buf = _build_xlsx(headers, clean_rows, sheet_name=Path(filename).stem[:31])
            output_files.append({
                "filename":     filename,
                "content":      primary_buf,
                "content_type": self.XLSX_CONTENT_TYPE,
            })
            messages.append(
                f"✅ **Extracted data** → `{filename}` "
                f"({record_count} records, {len(headers)} columns)"
            )
        else:
            # No structured extraction from AI — fall back to local parser
            fallback_files = self._local_extraction_fallback()
            output_files.extend(fallback_files)
            if fallback_files:
                names = ", ".join(f"`{f['filename']}`" for f in fallback_files)
                messages.append(f"✅ **Local extraction fallback** → {names}")

        # ── 4b. Conversion tool pipeline (ConversionService) ──────────
        conversion_files = self._run_conversion_pipeline()
        if conversion_files:
            output_files.extend(conversion_files)
            names = ", ".join(f"`{f['filename']}`" for f in conversion_files)
            messages.append(f"✅ **Conversion pipeline** → {names}")

        # ── 4c. Report xlsx ───────────────────────────────────────────
        # Always produce a report — it is the central deliverable.
        report_buf, report_filename = _build_report_xlsx(
            plan=plan,
            extraction=extraction if (headers and rows) else {},
            file_count=len(self.file_attachments),
        )
        output_files.append({
            "filename":     report_filename,
            "content":      report_buf,
            "content_type": self.XLSX_CONTENT_TYPE,
        })
        messages.append(f"📊 **Report** → `{report_filename}`")

        # ── 4d. DB-backed report via ReportService ────────────────────
        db_report_files = self._run_db_report(plan)
        if db_report_files:
            output_files.extend(db_report_files)
            names = ", ".join(f"`{f['filename']}`" for f in db_report_files)
            messages.append(f"📋 **DB Report** → {names}")

        # ── 4e. AI analysis ───────────────────────────────────────────
        if plan.get("analysis_requested"):
            analysis_text = self._run_ai_analysis(plan)
            if analysis_text:
                messages.append(f"🔍 **AI Analysis**\n{analysis_text}")

        # ── Build final response text ──────────────────────────────────
        steps_text = ""
        for step in plan.get("steps", []):
            steps_text += f"  {step.get('step', '?')}. `{step.get('tool', '?')}` — {step.get('reason', '')}\n"

        notes = plan.get("notes", "")
        response_parts = [
            f"**{plan.get('plan_summary', 'Processing complete.')}**\n",
        ]
        if steps_text:
            response_parts.append(f"**Steps executed:**\n{steps_text}")
        if summary:
            response_parts.append(f"**Extraction summary:** {summary}")
        response_parts.append("\n**Output files:**")
        response_parts.extend(f"- {m}" for m in messages)
        if notes:
            response_parts.append(f"\n**Notes:** {notes}")

        return output_files, "\n".join(response_parts)

    # ------------------------------------------------------------------
    # Sub-executors
    # ------------------------------------------------------------------

    def _local_extraction_fallback(self) -> list[dict]:
        """
        Use _extract_local_data_from_attachment from services as a fallback
        when the AI returned no structured extraction.
        """
        from .services import _build_output_files_from_attachments
        try:
            return _build_output_files_from_attachments(self.file_attachments)
        except Exception as exc:
            logger.warning("Local extraction fallback failed: %s", exc)
            return []

    def _run_conversion_pipeline(self) -> list[dict]:
        """
        Run ConversionService for each attachment.
        Returns output file dicts for any successful conversions.
        """
        from .services import _run_conversion_tools_for_attachments
        try:
            return _run_conversion_tools_for_attachments(
                self.file_attachments, self.user
            )
        except Exception as exc:
            logger.warning("Conversion pipeline failed: %s", exc)
            return []

    def _run_db_report(self, plan: dict) -> list[dict]:
        """
        Run ReportService for any conversion jobs that produced DB records.
        Detects report type from file extension.
        """
        output_files = []
        if not plan.get("report_requested", True):
            return output_files

        try:
            from tools.models import GeneratedReport
            from tools.services import ReportService
            from uploads.models import UploadBatch

            # Find the most recent batch created for this user's chat files
            batch = (
                UploadBatch.objects
                .filter(uploaded_by=self.user, notes__startswith="Chat attachment")
                .order_by("-created_at")
                .first()
            )
            if not batch:
                return output_files

            # Decide report type from file extensions in this batch
            ext_to_report = {
                "txt":  "ura_sales",
                "pdf":  "safaricom_dept",
                "xlsx": "acon_reconciliation",
                "xls":  "acon_reconciliation",
            }
            ext = (self.file_attachments[0].file_type or "").lower() if self.file_attachments else ""
            report_type = ext_to_report.get(ext, "ura_sales")

            report = ReportService.create_report(
                batch=batch,
                report_type=report_type,
                output_format="xlsx",
                generated_by=self.user,
            )
            ReportService.generate(report.id)
            report.refresh_from_db()

            if report.status == "ready" and report.file:
                report.file.open("rb")
                buf = BytesIO(report.file.read())
                output_files.append({
                    "filename":     Path(report.file.name).name,
                    "content":      buf,
                    "content_type": self.XLSX_CONTENT_TYPE,
                })

        except Exception as exc:
            logger.warning("DB report generation failed: %s", exc)

        return output_files

    def _run_ai_analysis(self, plan: dict) -> str:
        """
        Run AIEngineService.flag_anomalies (or summarise_batch) on the most
        recent batch associated with this user.  Returns a formatted string of
        insights, or an empty string on failure.
        """
        try:
            from ai_engine.services import AIEngineService
            from uploads.models import UploadBatch

            batch = (
                UploadBatch.objects
                .filter(uploaded_by=self.user, notes__startswith="Chat attachment")
                .order_by("-created_at")
                .first()
            )
            if not batch:
                return ""

            job = AIEngineService.create_job(
                batch=batch,
                task_type="flag_anomalies",
                user_prompt=f"Analyse the invoice data in batch {batch.pk} ({batch.period_label}).",
                requested_by=self.user,
            )
            AIEngineService.dispatch(job.id)
            job.refresh_from_db()

            insights = job.insights.order_by("-severity", "created_at")[:10]
            if not insights:
                return ""

            lines = []
            for ins in insights:
                icon = {"critical": "🔴", "warning": "🟡", "info": "🔵"}.get(ins.severity, "•")
                lines.append(f"{icon} **{ins.title}**: {ins.detail[:200]}")
            return "\n".join(lines)

        except Exception as exc:
            logger.warning("AI analysis step failed: %s", exc)
            return ""

    # ------------------------------------------------------------------
    # Fallback for no-file calls
    # ------------------------------------------------------------------

    def _no_files_response(self) -> str:
        return (
            "No files were attached. Please upload a URA fiscal receipt (.txt), "
            "a Safaricom bill (.pdf), or an ACON sales export (.xlsx / .xls) "
            "to begin processing."
        )