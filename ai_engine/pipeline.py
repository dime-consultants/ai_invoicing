# ai_engine/pipeline.py
import logging
import json
from io import BytesIO
from pathlib import Path

import openpyxl
import pdfplumber

from django.core.files.base import ContentFile
from .services import _get_client, GROK_MODEL
from .models import AIAnalysisJob, AIInsight

logger = logging.getLogger(__name__)


class DocumentPipeline:
    """LLM-powered unstructured document processing pipeline"""

    @staticmethod
    def _extract_raw_content(file_obj, filename: str) -> str:
        """Extract text from various file types"""
        ext = Path(filename).suffix.lower()
        content = ""

        try:
            if ext == '.txt':
                content = file_obj.read().decode('utf-8', errors='replace')

            elif ext == '.pdf':
                file_obj.seek(0)
                with pdfplumber.open(file_obj) as pdf:
                    for i, page in enumerate(pdf.pages, 1):
                        text = page.extract_text() or ""
                        content += f"\n--- Page {i} ---\n{text}\n"

            elif ext in ['.xlsx', '.xls']:
                file_obj.seek(0)
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    content += f"\n=== Sheet: {sheet_name} ===\n"
                    for row in ws.iter_rows(values_only=True):
                        content += str([str(cell) if cell is not None else "" for cell in row]) + "\n"

            elif ext == '.csv':
                content = file_obj.read().decode('utf-8', errors='replace')

            else:
                content = f"[Unsupported file type: {ext}]"

        except Exception as e:
            logger.exception(f"Failed to read {filename}")
            content = f"[ERROR READING FILE: {str(e)}]"

        return content[:120_000]  # Safe token limit

    @staticmethod
    def process_document(
        file_obj,
        filename: str,
        batch=None,
        requested_by=None,
        task_type: str = "custom",
        custom_prompt: str = ""
    ) -> AIAnalysisJob:
        """
        Main pipeline: File → AIAnalysisJob with structured output
        """
        raw_content = DocumentPipeline._extract_raw_content(file_obj, filename)

        system_prompt = """
You are an expert financial document parser for East African businesses (Uganda & Kenya).
Focus on URA fiscal receipts, credit notes, Safaricom invoices, ACON sales, etc.
Always respond with **valid JSON only**. No explanations outside the JSON.
"""

        user_prompt = f"""
File Name: {filename}

--- DOCUMENT START ---
{raw_content}
--- DOCUMENT END ---

Extract all key information into clean structured JSON.
Include a "records" array for individual transactions/receipts.

Preferred schema:
{{
  "document_type": "KN_PERIODICAL_REPORT | URA_FISCAL | SAFARICOM_INVOICE | ACON | OTHER",
  "company": "...",
  "period_from": "YYYY-MM-DD",
  "period_to": "YYYY-MM-DD",
  "total_records": 123,
  "total_amount": 1234567.89,
  "total_tax": 123456.78,
  "records": [
    {{
      "type": "FISCAL_RECEIPT | CREDIT_NOTE",
      "cu_invoice_number": "0040083410000140342",
      "date": "2026-04-10",
      "total_amount": 4862563.00,
      "tax_amount": 0.00
    }}
  ]
}}
"""

        if custom_prompt:
            user_prompt += f"\n\nAdditional Instructions:\n{custom_prompt}"

        # Create Job
        job = AIAnalysisJob.objects.create(
            batch=batch,
            task_type=task_type,
            user_prompt=user_prompt,
            system_prompt=system_prompt,
            requested_by=requested_by,
            status="queued"
        )

        # Run immediately (sync for now)
        DocumentPipeline._run_extraction_job(job)

        return job

    @staticmethod
    def _run_extraction_job(job: AIAnalysisJob):
        """Execute the LLM call"""
        try:
            client = _get_client()

            messages = [
                {"role": "system", "content": job.system_prompt},
                {"role": "user", "content": job.user_prompt}
            ]

            response = client.chat.completions.create(
                model=GROK_MODEL,
                messages=messages,
                temperature=0.0,
                max_tokens=8192,
            )

            raw_text = response.choices[0].message.content
            job.raw_response = raw_text
            job.input_tokens = response.usage.prompt_tokens
            job.output_tokens = response.usage.completion_tokens

            # Parse JSON
            structured = DocumentPipeline._safe_parse_json(raw_text)
            job.structured_output = structured
            job.status = "done"
            job.finished_at = datetime.now(timezone.utc)
            job.save()

            # Create Insights
            for rec in structured.get("records", [])[:20]:   # Limit insights
                AIInsight.objects.create(
                    job=job,
                    insight_type="summary_point",
                    severity="info",
                    reference_key=rec.get("cu_invoice_number", ""),
                    title=f"Record: {rec.get('cu_invoice_number', 'N/A')}",
                    detail=f"Total: {rec.get('total_amount')} | Tax: {rec.get('tax_amount')}"
                )

            logger.info(f"Document processing completed successfully: {job.pk}")

        except Exception as e:
            logger.exception("Document pipeline failed")
            job.status = "error"
            job.error_message = str(e)
            job.finished_at = datetime.now(timezone.utc)
            job.save()

    @staticmethod
    def _safe_parse_json(text: str):
        text = text.strip()
        if text.startswith("```"):
            text = text.split("```", 2)[-2].strip()
            if text.startswith("json"):
                text = text[4:].strip()
        try:
            return json.loads(text)
        except:
            return {"raw_text": text[:2000], "parse_error": True}