# chat/views.py
"""
Chat REST API views.

Endpoints
---------
GET  /api/chat/conversations/                       list user's conversations
POST /api/chat/conversations/                       create conversation
GET  /api/chat/conversations/<id>/                  detail (with messages)
PATCH|DELETE /api/chat/conversations/<id>/

GET  /api/chat/conversations/<id>/messages/         message list
POST /api/chat/conversations/<id>/send/             send message + optional files

GET  /api/chat/attachments/<id>/download/           download attachment

GET  /api/chat/workflows/                           list enabled workflows
GET  /api/chat/workflows/defaults/                  default sidebar workflows

POST /api/chat/message/                             stateless single-turn
GET  /api/chat/history/                             recent conversation messages
POST /api/chat/process-file/                        ingest + extract a file
POST /api/chat/convert-format/                      convert file format
POST /api/chat/export-data/                         serialise data to file

Changes vs. previous version
-----------------------------
- Removed __stateless__ sentinel conversations — stateless endpoints now
  create a real (non-sentinel) conversation so history isn't orphaned.
- Fixed history ordering (was inconsistent between WS and REST paths).
- Removed duplicate keyword-intent logic (moved fully into ChatService).
- ChatSimpleMessageView creates a proper conversation so attachments are
  always retrievable.
- _save_output_attachments is a module-level helper (used by chat/tasks.py).
"""
import base64
import csv
import json
import logging
from io import BytesIO, StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from django.core.files.base import ContentFile
from django.http import FileResponse
from django.views.generic import TemplateView

from rest_framework import generics, permissions, status, viewsets
from rest_framework.authentication import SessionAuthentication
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication

from .models import ChatConversation, ChatMessage, ChatMessageAttachment, Workflow
from .serializers import (
    ChatConversationSerializer,
    ChatConversationListSerializer,
    ChatMessageSerializer,
    ChatMessageAttachmentSerializer,
    WorkflowSerializer,
)
from .services import ChatService

logger = logging.getLogger(__name__)

MIME_MAP = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf":  "application/pdf",
    "csv":  "text/csv",
    "txt":  "text/plain",
    "json": "application/json",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


# ─────────────────────────────────────────────────────────────────────────────
# Module-level helpers (also imported by chat/tasks.py)
# ─────────────────────────────────────────────────────────────────────────────

def _save_output_attachments(output_files: list, assistant_msg) -> list:
    """
    Persist tool-produced output files as ChatMessageAttachment records.
    Returns the list of saved attachment instances.
    Called by both views (REST path) and chat/tasks.py (WS/Celery path).
    """
    saved = []
    for out in output_files:
        buf = out["content"]
        if hasattr(buf, "seek"):
            buf.seek(0)
        file_content = buf.read()
        if not file_content:
            logger.warning("Output file %s is empty — skipping.", out.get("filename"))
            continue

        ext = out["filename"].rsplit(".", 1)[-1].lower() if "." in out["filename"] else "bin"
        att = ChatMessageAttachment(
            message=assistant_msg,
            filename=out["filename"],
            file_type=ext,
            attachment_type="assistant_output",
            file_size_bytes=len(file_content),
        )
        att.file.save(out["filename"], ContentFile(file_content), save=True)
        saved.append(att)
    return saved


def _build_conv_history(conversation, exclude_pk=None) -> list[dict]:
    """Thin wrapper — see ChatService.build_conversation_history for the single
    source of truth shared with the WS consumer and the Celery chat task."""
    return ChatService.build_conversation_history(conversation, exclude_pk=exclude_pk)


def _parse_workflow_id(raw) -> tuple[int | None, str | None]:
    if not raw:
        return None, None
    raw = str(raw)
    if raw.startswith("full_"):
        return None, raw
    try:
        return int(raw), None
    except (TypeError, ValueError):
        return None, None


class ChatDispatchUnavailable(Exception):
    """The Celery broker could not be reached — dispatch() returned None."""


class ChatTurnTimeout(Exception):
    """The bounded wait elapsed before the worker finished the turn."""

    def __init__(self, turn_id: str):
        self.turn_id = turn_id
        super().__init__(f"Turn {turn_id} did not complete within the wait budget.")


def _run_turn_via_worker(
    *, conversation, user_msg, user, workflow_id=None, conversation_history=None,
):
    """
    Dispatch one chat turn to the Celery "chat" queue and block for a bounded
    time (settings.CHAT_SYNC_RESULT_TIMEOUT) waiting for the result, instead
    of running the agent inline in this request.

    Why: the actual LLM/tool-calling work (including any chunked map-reduce
    over a large file) can legitimately take longer than nginx's
    proxy_read_timeout on the HTTP path. Running it inline meant a slow turn
    either got killed mid-response (no answer, connection dropped) or, for
    turns just under the wire, tied up this request's thread with as many
    parallel outbound calls as the turn needed — with no system-wide cap on
    how many such bursts could run at once. Running it in the worker instead
    means: the task has a real time budget (8 min soft limit, not ~60s), and
    the worker pool's own autoscale setting is the one place total concurrent
    LLM calls is bounded, rather than every request improvising its own.

    Returns (assistant_msg, output_attachments) — the objects
    run_chat_turn_task already created and saved, not new ones, so callers
    serialize them the same way regardless of how the turn ran.

    Raises ChatDispatchUnavailable if the broker can't be reached at all
    (nothing was queued), or ChatTurnTimeout if the wait budget elapses.
    Neither means the turn was lost: a timed-out turn keeps running in the
    worker and will still save its result — the caller just didn't get it
    back in time for *this* response. Any other exception raised by the task
    itself (a real failure) propagates to the caller unchanged.
    """
    import uuid

    from celery.exceptions import TimeoutError as CeleryTimeoutError
    from django.conf import settings

    from chat.tasks import run_chat_turn_task
    from config.dispatch import dispatch

    turn_id = uuid.uuid4().hex
    async_result = dispatch(
        run_chat_turn_task,
        turn_id=turn_id,
        conversation_id=conversation.pk,
        user_message_id=user_msg.pk,
        user_id=user.pk,
        workflow_id=workflow_id,
        conversation_history=conversation_history or [],
    )
    if async_result is None:
        raise ChatDispatchUnavailable(
            "Could not queue the message — the task broker is unavailable."
        )

    timeout = getattr(settings, "CHAT_SYNC_RESULT_TIMEOUT", 45)
    try:
        result = async_result.get(timeout=timeout)
    except CeleryTimeoutError:
        raise ChatTurnTimeout(turn_id)

    assistant_message_id = (result or {}).get("assistant_message_id")
    if not assistant_message_id:
        error = (result or {}).get("error") or "The chat turn did not complete."
        raise RuntimeError(error)

    assistant_msg = ChatMessage.objects.get(pk=assistant_message_id)
    output_attachments = list(
        ChatMessageAttachment.objects.filter(pk__in=result.get("attachment_ids") or [])
    )
    return assistant_msg, output_attachments


def _get_or_create_working_conversation(user, title: str = "Chat") -> ChatConversation:
    """
    Return the user's most recent active conversation or create a new one.
    Used by stateless endpoints so they don't pollute the DB with sentinel rows.
    """
    conv = (
        ChatConversation.objects
        .filter(user=user, is_active=True)
        .exclude(title__startswith="__")
        .order_by("-updated_at")
        .first()
    )
    if conv:
        return conv
    return ChatConversation.objects.create(user=user, title=title)


# ─────────────────────────────────────────────────────────────────────────────
# UI entry point
# ─────────────────────────────────────────────────────────────────────────────

class ChatInterfaceView(TemplateView):
    template_name = "chat/interface.html"


# ─────────────────────────────────────────────────────────────────────────────
# Conversations
# ─────────────────────────────────────────────────────────────────────────────

class ChatConversationListCreateView(generics.ListCreateAPIView):
    """
    GET  /api/chat/conversations/  — list (lightweight, no messages)
    POST /api/chat/conversations/  — create
    """
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        return ChatConversationListSerializer if self.request.method == "GET" else ChatConversationSerializer

    def get_queryset(self):
        return (
            ChatConversation.objects
            .filter(user=self.request.user)
            .exclude(title__startswith="__")
            .order_by("-updated_at")
        )

    def create(self, request, *args, **kwargs):
        title = (
            (request.data.get("title") or "").strip()
            if isinstance(request.data, dict)
            else ""
        ) or "Untitled Conversation"
        conv = ChatConversation.objects.create(user=request.user, title=title)
        return Response(ChatConversationSerializer(conv).data, status=status.HTTP_201_CREATED)


class ChatConversationDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET    /api/chat/conversations/<id>/
    PATCH  /api/chat/conversations/<id>/  — title / is_active only
    DELETE /api/chat/conversations/<id>/
    """
    permission_classes = [permissions.IsAuthenticated]
    serializer_class   = ChatConversationSerializer

    def get_queryset(self):
        return ChatConversation.objects.filter(user=self.request.user)

    def update(self, request, *args, **kwargs):
        allowed   = {"title", "is_active"}
        data      = request.data if isinstance(request.data, dict) else {}
        updates   = {k: v for k, v in data.items() if k in allowed}
        instance  = self.get_object()
        for field, value in updates.items():
            setattr(instance, field, value)
        if updates:
            instance.save(update_fields=list(updates.keys()) + ["updated_at"])
        return Response(ChatConversationSerializer(instance).data)


# ─────────────────────────────────────────────────────────────────────────────
# Send message
# ─────────────────────────────────────────────────────────────────────────────

class ChatMessageSendView(APIView):
    """
    POST /api/chat/conversations/<conversation_id>/send/

    multipart/form-data:
        message      str       — required (or at least one file)
        files        file[]    — optional attachments
        workflow_id  int|str   — optional workflow selector

    Returns (default, synchronous):
        201 — user_message, assistant_message, user_attachments, output_attachments

    Returns (async, opt-in via ?sync=0 / ?async=1):
        202 — status, turn_id, conversation_id, user_message. The assistant reply
              is pushed over /ws/chat/<conversation_id>/; callers that cannot
              listen there must use the default synchronous path.
    """
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    permission_classes     = [permissions.IsAuthenticated]

    def post(self, request, conversation_id):
        try:
            conversation = ChatConversation.objects.get(
                pk=conversation_id, user=request.user
            )
        except ChatConversation.DoesNotExist:
            return Response(
                {"error": "Conversation not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        user_input     = (request.data.get("message") or "").strip()
        uploaded_files = request.FILES.getlist("files")

        if not user_input and not uploaded_files:
            return Response(
                {"error": "Provide a message or at least one file."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not user_input:
            user_input = "Please extract all data from the attached file and produce an Excel file."

        # Save user message
        user_msg = ChatMessage.objects.create(
            conversation=conversation, role="user", content=user_input,
        )

        # Persist uploaded files as ChatMessageAttachment records
        user_attachments = []
        for f in uploaded_files:
            ext = f.name.rsplit(".", 1)[-1].lower() if "." in f.name else "other"
            att = ChatMessageAttachment.objects.create(
                message=user_msg,
                file=f,
                filename=f.name,
                file_type=ext,
                attachment_type="user_upload",
                file_size_bytes=f.size,
            )
            user_attachments.append(att)

        workflow_id_int, _ = _parse_workflow_id(request.data.get("workflow_id"))
        conversation_history = _build_conv_history(conversation, exclude_pk=user_msg.pk)

        # Synchronous is the default: REST callers read assistant_message straight
        # off this response and have no WebSocket/polling fallback. The async path
        # (202 + turn_id, result pushed over the conversation WS) is opt-in via
        # ?sync=0 or ?async=1 — only use it from a client that listens on the WS.
        sync_flag  = request.query_params.get("sync")
        if sync_flag is None:
            sync_flag = request.data.get("sync")
        async_flag = request.query_params.get("async")
        if async_flag is None:
            async_flag = request.data.get("async")

        if sync_flag is not None:
            sync = str(sync_flag).lower() in ("1", "true", "yes")
        elif async_flag is not None:
            sync = str(async_flag).lower() not in ("1", "true", "yes")
        else:
            sync = True

        if sync:
            try:
                assistant_msg, output_attachments = _run_turn_via_worker(
                    conversation=conversation,
                    user_msg=user_msg,
                    user=request.user,
                    workflow_id=workflow_id_int,
                    conversation_history=conversation_history,
                )
            except ChatDispatchUnavailable as exc:
                return Response(
                    {"error": str(exc),
                     "user_message": ChatMessageSerializer(user_msg, context={"request": request}).data},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE,
                )
            except ChatTurnTimeout as exc:
                # The turn is NOT lost — it keeps running in the worker and
                # will still save its result. This only means it didn't
                # finish within the wait budget for this response; the
                # frontend's existing error-toast path (any non-2xx with an
                # "error" field) surfaces this without needing to understand
                # turn_id/polling.
                return Response(
                    {"error": "This is taking longer than usual (likely a large file). "
                              "It's still processing — please check back in a moment or "
                              "send your message again.",
                     "turn_id": exc.turn_id,
                     "user_message": ChatMessageSerializer(user_msg, context={"request": request}).data},
                    status=status.HTTP_504_GATEWAY_TIMEOUT,
                )
            except Exception as exc:
                logger.exception("Chat turn failed: %s", exc)
                return Response(
                    {"error": f"Failed to generate response: {exc}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            ctx = {"request": request}
            return Response(
                {
                    "user_message":       ChatMessageSerializer(user_msg, context=ctx).data,
                    "assistant_message":  ChatMessageSerializer(assistant_msg, context=ctx).data,
                    "user_attachments":   ChatMessageAttachmentSerializer(user_attachments, many=True, context=ctx).data,
                    "output_attachments": ChatMessageAttachmentSerializer(output_attachments, many=True, context=ctx).data,
                },
                status=status.HTTP_201_CREATED,
            )

        import uuid
        from chat.tasks import run_chat_turn_task
        from config.dispatch import dispatch

        turn_id = uuid.uuid4().hex
        queued = dispatch(
            run_chat_turn_task,
            turn_id=turn_id,
            conversation_id=conversation.pk,
            user_message_id=user_msg.pk,
            user_id=request.user.pk,
            workflow_id=workflow_id_int,
            conversation_history=conversation_history,
        )
        if queued is None:
            # user_msg is already committed; without a queued turn nothing will
            # ever answer it. Say so instead of 500-ing on a kombu error.
            return Response(
                {"error": "Could not queue the message — the task broker is "
                          "unavailable. Retry, or resend with ?sync=1.",
                 "user_message": ChatMessageSerializer(user_msg, context={"request": request}).data},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        ctx = {"request": request}
        return Response(
            {
                "status": "accepted",
                "turn_id": turn_id,
                "conversation_id": conversation.pk,
                "user_message": ChatMessageSerializer(user_msg, context=ctx).data,
                "message": "Turn accepted. Listen on the conversation WebSocket or poll messages.",
            },
            status=status.HTTP_202_ACCEPTED,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Message list
# ─────────────────────────────────────────────────────────────────────────────

class ChatMessageListView(generics.ListAPIView):
    """GET /api/chat/conversations/<conversation_id>/messages/"""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class   = ChatMessageSerializer

    def get_queryset(self):
        try:
            conv = ChatConversation.objects.get(
                pk=self.kwargs["conversation_id"], user=self.request.user
            )
        except ChatConversation.DoesNotExist:
            return ChatMessage.objects.none()
        return (
            ChatMessage.objects
            .filter(conversation=conv)
            .exclude(role="system")
            .order_by("created_at", "pk")
            .prefetch_related("attachments")
        )

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx


# ─────────────────────────────────────────────────────────────────────────────
# Attachment download
# ─────────────────────────────────────────────────────────────────────────────

class ChatAttachmentDownloadView(APIView):
    """GET /api/chat/attachments/<id>/download/"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, attachment_id):
        try:
            att = ChatMessageAttachment.objects.select_related(
                "message__conversation"
            ).get(pk=attachment_id)
        except ChatMessageAttachment.DoesNotExist:
            return Response({"error": "Attachment not found."}, status=status.HTTP_404_NOT_FOUND)

        if att.message.conversation.user != request.user:
            return Response({"error": "Access denied."}, status=status.HTTP_403_FORBIDDEN)

        if not att.file:
            return Response({"error": "File not available."}, status=status.HTTP_404_NOT_FOUND)

        ct = MIME_MAP.get(att.file_type, "application/octet-stream")
        resp = FileResponse(att.file.open("rb"), content_type=ct, as_attachment=True)
        resp["Content-Disposition"] = f'attachment; filename="{att.filename}"'
        return resp


# ─────────────────────────────────────────────────────────────────────────────
# Workflows
# ─────────────────────────────────────────────────────────────────────────────

class WorkflowViewSet(viewsets.ReadOnlyModelViewSet):
    """
    GET /api/chat/workflows/           all enabled workflows
    GET /api/chat/workflows/defaults/  default sidebar workflows
    GET /api/chat/workflows/?type=...  filter by workflow_type
    """
    serializer_class   = WorkflowSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = Workflow.objects.filter(enabled=True)
        if wf_type := self.request.query_params.get("type"):
            qs = qs.filter(workflow_type=wf_type)
        return qs

    @action(detail=False, methods=["get"])
    def defaults(self, request):
        qs = Workflow.objects.filter(enabled=True, is_default=True)
        return Response(WorkflowSerializer(qs, many=True).data)


# ─────────────────────────────────────────────────────────────────────────────
# Stateless single-turn  (POST /api/chat/message/)
# ─────────────────────────────────────────────────────────────────────────────

class ChatSimpleMessageView(APIView):
    """
    POST /api/chat/message/

    Stateless single-turn. Creates a real conversation (or reuses the most
    recent one) — no sentinel __stateless__ rows.

    multipart/form-data or JSON:
        message               str
        files                 file[]
        conversation_history  JSON string (optional)
        workflow_id           int (optional)
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser]

    def post(self, request):
        message  = (request.data.get("message") or "").strip()

        uploaded_files = request.FILES.getlist("files") or []
        if not message and not uploaded_files:
            return Response(
                {"error": {"code": "bad_request",
                           "message": "Provide a message or files.", "details": None}},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not message:
            message = "Please extract all data from the attached file and produce an Excel file."

        workflow_id_int, _ = _parse_workflow_id(request.data.get("workflow_id"))

        # Use a real conversation so attachments are retrievable. Once a real
        # conversation exists, its DB-derived history is authoritative — a
        # client-supplied conversation_history payload could be stale or
        # simply wrong, so it is never allowed to override it here.
        conv = _get_or_create_working_conversation(request.user, title="Quick Chat")
        user_msg = ChatMessage.objects.create(
            conversation=conv, role="user", content=message,
        )

        temp_attachments = []
        for f in uploaded_files:
            ext = f.name.rsplit(".", 1)[-1].lower() if "." in f.name else "other"
            att = ChatMessageAttachment.objects.create(
                message=user_msg,
                file=f, filename=f.name, file_type=ext,
                attachment_type="user_upload", file_size_bytes=f.size,
            )
            temp_attachments.append(att)

        try:
            assistant_msg, output_attachments = _run_turn_via_worker(
                conversation=conv,
                user_msg=user_msg,
                user=request.user,
                workflow_id=workflow_id_int,
                conversation_history=_build_conv_history(conv, user_msg.pk),
            )
        except ChatDispatchUnavailable as exc:
            return Response(
                {"error": {"code": "service_unavailable", "message": str(exc), "details": None}},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except ChatTurnTimeout as exc:
            # Not lost — still running in the worker, just not back in time
            # for this response. See ChatMessageSendView for the same pattern.
            return Response(
                {"error": {
                    "code": "timeout",
                    "message": "This is taking longer than usual (likely a large file). "
                               "It's still processing — please check back in a moment or "
                               "send your message again.",
                    "details": {"turn_id": exc.turn_id},
                }},
                status=status.HTTP_504_GATEWAY_TIMEOUT,
            )
        except Exception as exc:
            logger.exception("ChatSimpleMessageView failed: %s", exc)
            return Response(
                {"error": {"code": "server_error", "message": str(exc), "details": None}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        processed_data = None
        if output_attachments:
            first = output_attachments[0]
            ext = first.filename.rsplit(".", 1)[-1] if "." in first.filename else "xlsx"
            first.file.open("rb")
            try:
                content_bytes = first.file.read()
            finally:
                first.file.close()
            processed_data = {
                "format":   ext,
                "filename": first.filename,
                "content":  base64.b64encode(content_bytes).decode("utf-8"),
            }

        return Response({"response": assistant_msg.content, "processed_data": processed_data})


# ─────────────────────────────────────────────────────────────────────────────
# Chat history
# ─────────────────────────────────────────────────────────────────────────────

class ChatHistoryView(APIView):
    """GET /api/chat/history?conversationId=<id>&limit=50"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        conversation_id = request.query_params.get("conversationId")
        try:
            limit = int(request.query_params.get("limit", 50))
        except (TypeError, ValueError):
            limit = 50          # ?limit=abc used to be an unhandled 500
        limit = max(1, min(limit, 200))

        if conversation_id:
            try:
                conv = ChatConversation.objects.get(
                    pk=conversation_id, user=request.user
                )
            except ChatConversation.DoesNotExist:
                return Response(
                    {"error": {"code": "not_found",
                               "message": "Conversation not found.", "details": None}},
                    status=status.HTTP_404_NOT_FOUND,
                )
        else:
            conv = (
                ChatConversation.objects
                .filter(user=request.user, is_active=True)
                .exclude(title__startswith="__")
                .order_by("-updated_at")
                .first()
            )
            if not conv:
                return Response({"messages": []})

        messages = (
            ChatMessage.objects
            .filter(conversation=conv)
            .exclude(role="system")
            .order_by("created_at", "pk")
            .prefetch_related("attachments")
        )[:limit]

        ctx = {"request": request}
        return Response({"messages": ChatMessageSerializer(messages, many=True, context=ctx).data})


# ─────────────────────────────────────────────────────────────────────────────
# Process file
# ─────────────────────────────────────────────────────────────────────────────

class ChatProcessFileView(APIView):
    """
    POST /api/chat/process-file/
    multipart/form-data: file (required), output_format (xlsx|csv|json|txt)
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response(
                {"error": {"code": "bad_request", "message": "file is required.", "details": None}},
                status=status.HTTP_400_BAD_REQUEST,
            )
        output_format = (request.data.get("output_format") or "xlsx").lower()

        conv     = _get_or_create_working_conversation(request.user, title="File Processing")
        user_msg = ChatMessage.objects.create(
            conversation=conv, role="user",
            content=f"Extract data from {uploaded.name}",
        )
        ext = uploaded.name.rsplit(".", 1)[-1].lower() if "." in uploaded.name else "other"
        # Not bound to a name — run_chat_turn_task discovers it itself via
        # ChatMessageAttachment.objects.filter(message=user_msg).
        ChatMessageAttachment.objects.create(
            message=user_msg, file=uploaded, filename=uploaded.name,
            file_type=ext, attachment_type="user_upload", file_size_bytes=uploaded.size,
        )

        content = b""
        out_filename = Path(uploaded.name).stem + f"_processed.{output_format}"
        warnings = []

        try:
            _, output_attachments = _run_turn_via_worker(
                conversation=conv,
                user_msg=user_msg,
                user=request.user,
                conversation_history=_build_conv_history(conv, user_msg.pk),
            )
            if output_attachments:
                first = output_attachments[0]
                first.file.open("rb")
                try:
                    content = first.file.read()
                finally:
                    first.file.close()
                out_filename = first.filename
            else:
                warnings.append("No output file produced.")
        except ChatDispatchUnavailable as exc:
            warnings.append(str(exc))
        except ChatTurnTimeout:
            # Not lost — still running in the worker, just not back in time
            # for this response.
            warnings.append(
                "This is taking longer than usual (likely a large file). It's "
                "still processing — please check back in a moment or retry."
            )
        except Exception as exc:
            logger.exception("ChatProcessFileView failed: %s", exc)
            warnings.append(str(exc))

        return Response({
            "status": "success" if not warnings else "warning",
            "processed_file": {
                "filename": out_filename,
                "format":   output_format,
                "size":     len(content),
                "content":  base64.b64encode(content).decode("utf-8") if content else "",
            },
            "summary": {
                "rows_processed": content.count(b"\n") + 1 if content else 0,
                "errors":         0,
                "warnings":       warnings,
            },
        })


# ─────────────────────────────────────────────────────────────────────────────
# Convert format
# ─────────────────────────────────────────────────────────────────────────────

class ChatConvertFormatView(APIView):
    """
    POST /api/chat/convert-format/
    multipart/form-data: file (required), to_format (xlsx|csv|json|txt)
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded  = request.FILES.get("file")
        to_format = (request.data.get("to_format") or "xlsx").lower()
        if not uploaded:
            return Response(
                {"error": {"code": "bad_request", "message": "file is required.", "details": None}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        conv     = _get_or_create_working_conversation(request.user, title="Format Conversion")
        user_msg = ChatMessage.objects.create(
            conversation=conv, role="user",
            content=f"Convert {uploaded.name} to {to_format}",
        )
        ext = uploaded.name.rsplit(".", 1)[-1].lower() if "." in uploaded.name else "other"
        # Not bound to a name — run_chat_turn_task discovers it itself via
        # ChatMessageAttachment.objects.filter(message=user_msg).
        ChatMessageAttachment.objects.create(
            message=user_msg, file=uploaded, filename=uploaded.name,
            file_type=ext, attachment_type="user_upload", file_size_bytes=uploaded.size,
        )

        try:
            _, output_attachments = _run_turn_via_worker(
                conversation=conv,
                user_msg=user_msg,
                user=request.user,
                conversation_history=_build_conv_history(conv, user_msg.pk),
            )
            if not output_attachments:
                return Response(
                    {"error": {"code": "server_error",
                               "message": "Conversion produced no output.", "details": None}},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            first = output_attachments[0]
            first.file.open("rb")
            try:
                content = first.file.read()
            finally:
                first.file.close()
            out_name = first.filename
        except ChatDispatchUnavailable as exc:
            return Response(
                {"error": {"code": "service_unavailable", "message": str(exc), "details": None}},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except ChatTurnTimeout as exc:
            # Not lost — still running in the worker, just not back in time
            # for this response.
            return Response(
                {"error": {
                    "code": "timeout",
                    "message": "This is taking longer than usual (likely a large file). "
                               "It's still processing — please check back in a moment or retry.",
                    "details": {"turn_id": exc.turn_id},
                }},
                status=status.HTTP_504_GATEWAY_TIMEOUT,
            )
        except Exception as exc:
            logger.exception("ChatConvertFormatView: %s", exc)
            return Response(
                {"error": {"code": "server_error", "message": str(exc), "details": None}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({
            "filename":  out_name,
            "format":    to_format,
            "content":   base64.b64encode(content).decode("utf-8"),
            "mime_type": MIME_MAP.get(to_format, "application/octet-stream"),
        })


# ─────────────────────────────────────────────────────────────────────────────
# Export data
# ─────────────────────────────────────────────────────────────────────────────

class ChatExportDataView(APIView):
    """
    POST /api/chat/export-data/
    JSON body: { "data": [...], "format": "csv|json|xlsx|txt",
                 "filename": "output", "options": {...} }
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        data     = request.data.get("data")
        fmt      = (request.data.get("format") or "xlsx").lower()
        filename = (request.data.get("filename") or "export").rstrip(".")
        options  = request.data.get("options") or {}
        include_hdr = options.get("include_headers", True)
        delimiter   = options.get("delimiter", ",")

        if data is None:
            return Response(
                {"error": {"code": "bad_request", "message": "data is required.", "details": None}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        out_name = f"{filename}.{fmt}"
        content  = b""

        try:
            rows = data if isinstance(data, list) else [data]

            if fmt == "json":
                content = json.dumps(data, indent=2, default=str).encode("utf-8")

            elif fmt in ("csv", "txt"):
                sep = delimiter if fmt == "csv" else "\t"
                buf = StringIO()
                writer = csv.writer(buf, delimiter=sep)
                if rows and isinstance(rows[0], dict):
                    if include_hdr:
                        writer.writerow(rows[0].keys())
                    for row in rows:
                        writer.writerow(row.values())
                elif rows and isinstance(rows[0], list):
                    for row in rows:
                        writer.writerow(row)
                else:
                    for row in rows:
                        writer.writerow([str(row)])
                content = buf.getvalue().encode("utf-8")

            elif fmt == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Export"
                start_row = 1
                if rows and isinstance(rows[0], dict) and include_hdr:
                    headers = list(rows[0].keys())
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=str(h))
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="1F4E79")
                    start_row = 2
                for r_idx, row in enumerate(rows, start_row):
                    if isinstance(row, dict):
                        for c_idx, val in enumerate(row.values(), 1):
                            ws.cell(row=r_idx, column=c_idx, value=val)
                    elif isinstance(row, list):
                        for c_idx, val in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=val)
                    else:
                        ws.cell(row=r_idx, column=1, value=str(row))
                buf = BytesIO()
                wb.save(buf)
                content = buf.getvalue()
            else:
                content = json.dumps(data, indent=2, default=str).encode("utf-8")

        except Exception as exc:
            logger.exception("ChatExportDataView: %s", exc)
            return Response(
                {"error": {"code": "server_error", "message": str(exc), "details": None}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({
            "filename":  out_name,
            "format":    fmt,
            "content":   base64.b64encode(content).decode("utf-8"),
            "mime_type": MIME_MAP.get(fmt, "application/octet-stream"),
        })