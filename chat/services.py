# chat/services.py
import csv
import json
import logging
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone

from uploads.models import UploadBatch
from uploads.services import UploadService
from .models import Workflow

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Intent detection
# ─────────────────────────────────────────────────────────────────────────────

_INTENT_KEYWORDS = {
    "help":        ["help", "what can you do", "capabilities", "guide", "how to"],
    "upload":      ["upload", "add file", "new batch", "import", "ingest"],
    "process":     ["process", "parse", "convert", "execute", "reparse"],
    "reconcile":   ["reconcile", "reconciliation", "match", "variance", "ura vs", "acon"],
    "report":      ["report", "generate", "export", "download"],
    "status":      ["status", "check", "progress", "pending", "failed"],
    "ai_analysis": ["analyse", "analyze", "anomaly", "flag", "summarise", "summarize"],
    "create_file": ["create", "make", "build", "extract", "invoice", "spreadsheet", "xlsx", "excel"],
}

_INTENT_SYSTEM_PROMPTS = {
    "help": (
        "You are an AI assistant in an invoice processing tool for Kuehne + Nagel's "
        "finance team. Explain what the tool does and suggest next steps. Cover: "
        "URA fiscal .txt files, Safaricom PDF bills, ACON .xlsx exports, reconciliation, "
        "AI anomaly detection, and Excel report generation."
    ),
    "upload": (
        "You are an AI assistant in an invoice processing tool. Guide the user through "
        "creating a batch and uploading files. Be specific about file types and endpoints."
    ),
    "process": (
        "You are an AI assistant in an invoice processing tool. Explain how files are "
        "parsed automatically after upload and how to use the reparse endpoint for failures."
    ),
    "reconcile": (
        "You are an AI assistant in an invoice processing tool. Explain the three "
        "reconciliation types: ura_vs_acon, safaricom_invoice, acon_variance."
    ),
    "report": (
        "You are an AI assistant in an invoice processing tool. Explain available "
        "report types and how to download them."
    ),
    "status": (
        "You are an AI assistant in an invoice processing tool. Live batch data is "
        "included in the message. Summarise each batch's state and suggest next actions."
    ),
    "ai_analysis": (
        "You are an AI assistant in an invoice processing tool. Explain AI job types: "
        "summarise_batch, flag_anomalies, explain_variance, classify_lines, custom."
    ),
    "create_file": (
        "You are a financial data analyst. The user has attached a file. "
        "Read its content carefully and extract ALL data rows into a structured format.\n\n"
        "Available parsing tools are: txt_to_xlsx, pdf_to_xlsx, xlsx_clean. "
        "Select the appropriate tool for each file type and explain which one was used.\n\n"
        "For URA fiscal receipt .txt files, extract each receipt block:\n"
        "  Headers: CU Invoice Number, Date, Time, Total (UGX), Taxes (UGX), Entry Type\n"
        "  Parse every block that starts with 'FISCAL RECEIPT' or 'CREDIT NOTE'.\n"
        "  Numbers contain spaces as thousands separators and commas as decimal — normalise them.\n\n"
        "For Safaricom bills extract: Name, Reference NO, Invoice NO, Net Amount, VAT, Excise, Billed Amount.\n"
        "For ACON exports extract all columns as-is.\n\n"
        "Respond with ONLY a JSON object — no preamble, no markdown fences:\n"
        "{\n"
        '  "tool": "txt_to_xlsx|pdf_to_xlsx|xlsx_clean",\n'
        '  "headers": ["Col1", "Col2", ...],\n'
        '  "rows": [["val1", "val2"], ...],\n'
        '  "filename": "descriptive_snake_case_name.xlsx",\n'
        '  "summary": "2-3 sentence description of what was extracted",\n'
        '  "notes": "Optional notes on why this tool was selected"\n'
        "}\n"
        "If the tool is selected, prefer tool-based conversion rather than raw model-only extraction."
        "Extract ALL rows. Do not summarise or truncate the data."
    ),
}

_TOOL_DEFINITIONS = {
    "txt_to_xlsx": {
        "description": "Parse URA fiscal receipt .txt files and produce an Excel file with receipt rows.",
        "file_types": ["txt"],
        "output": "Excel (.xlsx) with headers: CU Invoice Number, Date, Time, Total, Tax Amount, Entry Type",
    },
    "pdf_to_xlsx": {
        "description": "Parse Safaricom invoice PDF tables and produce an Excel file with line item rows.",
        "file_types": ["pdf"],
        "output": "Excel (.xlsx) with headers: Name, Reference NO., Invoice NO., Net Amount, VAT, Excise, Billed Amount",
    },
    "xlsx_clean": {
        "description": "Read attached spreadsheet files and normalize rows for downstream analysis.",
        "file_types": ["xlsx", "xls"],
        "output": "Cleaned Excel (.xlsx) with the same columns and normalized values.",
    },
}

_GENERAL_SYSTEM_PROMPT = (
    "You are a knowledgeable AI assistant in an invoice processing tool used by "
    "Kuehne + Nagel's finance team. Answer the user's question accurately and practically."
)


def _detect_intent(message: str, has_files: bool = False) -> str:
    if has_files:
        lower = message.lower()
        if any(kw in lower for kw in _INTENT_KEYWORDS["create_file"]):
            return "create_file"
        return "create_file"   # any attachment defaults to file extraction
    lower = message.lower()
    for intent, keywords in _INTENT_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            return intent
    return "general"


def _get_status_context() -> str:
    batches = UploadBatch.objects.order_by("-created_at")[:10]
    if not batches:
        return "No batches exist yet."
    lines = ["Recent batches:"]
    for b in batches:
        fc = b.files.count()
        ec = b.files.filter(parse_status="parse_error").count()
        lines.append(
            f"  Batch {b.pk} | {b.label} | "
            f"status={b.status} | files={fc} | errors={ec}"
        )
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# File reading
# ─────────────────────────────────────────────────────────────────────────────

def _read_file_to_text(file_obj, filename: str, max_lines: int = 2000) -> str:
    """
    Convert any supported file to a plain-text representation for Grok.
    Truncates at max_lines to stay within token limits.
    """
    ext = Path(filename).suffix.lower()
    try:
        if ext == ".txt":
            file_obj.seek(0)
            raw = file_obj.read().decode("utf-8", errors="replace")
            lines = raw.splitlines()
            truncated = len(lines) > max_lines
            preview = "\n".join(lines[:max_lines])
            suffix = f"\n\n[Truncated — {len(lines)} total lines in file]" if truncated else ""
            return f"[TXT FILE: {filename}]\n\n{preview}{suffix}"

        if ext == ".csv":
            file_obj.seek(0)
            raw = file_obj.read().decode("utf-8", errors="replace")
            rows = list(csv.reader(StringIO(raw)))
            truncated = len(rows) > max_lines
            lines = [",".join(str(c) for c in r) for r in rows[:max_lines]]
            suffix = f"\n[Truncated — {len(rows)} total rows]" if truncated else ""
            return f"[CSV FILE: {filename}]\n\n" + "\n".join(lines) + suffix

        if ext in (".xlsx", ".xls"):
            file_obj.seek(0)
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            parts = [f"[EXCEL FILE: {filename}]"]
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"\n=== Sheet: {sheet_name} ===")
                count = 0
                for row in ws.iter_rows(values_only=True):
                    if not any(c is not None for c in row):
                        continue
                    parts.append("\t".join("" if c is None else str(c) for c in row))
                    count += 1
                    if count >= max_lines:
                        parts.append(f"[Truncated — {ws.max_row} total rows]")
                        break
            return "\n".join(parts)

        if ext == ".pdf":
            try:
                import pdfplumber
                file_obj.seek(0)
                parts = [f"[PDF FILE: {filename}]"]
                with pdfplumber.open(file_obj) as pdf:
                    total_rows = 0
                    for i, page in enumerate(pdf.pages[:10], 1):
                        parts.append(f"\n--- Page {i} ---")
                        text = page.extract_text() or ""
                        if text:
                            parts.append(text)
                        for table in (page.extract_tables() or []):
                            parts.append("[TABLE]")
                            for row in table:
                                parts.append("\t".join("" if c is None else str(c) for c in row))
                                total_rows += 1
                                if total_rows >= max_lines:
                                    parts.append("[Truncated]")
                                    break
                            if total_rows >= max_lines:
                                break
                        if total_rows >= max_lines:
                            break
                    if len(pdf.pages) > 10:
                        parts.append(f"[{len(pdf.pages) - 10} more pages not shown]")
                return "\n".join(parts)
            except ImportError:
                return f"[PDF FILE: {filename}] — pdfplumber not installed"

        return f"[UNSUPPORTED FORMAT: {ext}]"

    except Exception as exc:
        logger.exception("Error reading %s: %s", filename, exc)
        return f"[ERROR READING {filename}: {exc}]"


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_xlsx(headers: list, rows: list, sheet_name: str = "Data") -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if value is not None and str(value).strip():
                cleaned = str(value).replace(",", "").replace(" ", "")
                try:
                    cell.value = float(cleaned)
                except (ValueError, InvalidOperation):
                    cell.value = str(value)
            else:
                cell.value = str(value) if value is not None else ""
            cell.alignment = Alignment(horizontal="left", vertical="top")

    for col_idx, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = len(str(header))
        for row_idx in range(2, min(len(rows) + 2, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Grok call
# ─────────────────────────────────────────────────────────────────────────────

def generate_chat_response(
    user_message: str,
    system_prompt: str = "",
    conversation_history: list[dict] | None = None,
) -> str:
    from ai_engine.services import _get_client, GROK_MODEL

    sp = system_prompt if isinstance(system_prompt, str) and system_prompt else _GENERAL_SYSTEM_PROMPT
    messages = [{"role": "system", "content": sp}]

    if conversation_history:
        for entry in conversation_history:
            if isinstance(entry, dict):
                role = entry.get("role", "")
                content = entry.get("content", "")
                if isinstance(role, str) and isinstance(content, str):
                    messages.append({"role": role, "content": content})

    messages.append({"role": "user", "content": str(user_message)})

    # DEBUG — remove after fix
    #import json as _json
    #for i, m in enumerate(messages):
    #    try:
    #        _json.dumps(m)
    #    except TypeError as e:
    #        logger.error("NON-SERIALIZABLE at messages[%d]: %s | value types: %s",
    #                    i, e, {k: type(v).__name__ for k, v in m.items()})

    response = _get_client().chat.completions.create(
        model=GROK_MODEL(),
        messages=messages,
        temperature=0.2,
        max_tokens=4096,
    )
    return response.choices[0].message.content.strip()


def _extract_local_data_from_attachment(att):
    """Attempt local structured extraction from an attachment."""
    ext = Path(att.filename).suffix.lower()

    def _parse_txt_invoice_blocks(text: str) -> tuple[list[str], list[list[str]]]:
        blocks = re.split(r'(?=FISCAL RECEIPT|CREDIT NOTE)', text, flags=re.IGNORECASE)
        rows = []
        for block in blocks:
            if not block.strip():
                continue
            block_text = block.strip()
            if not re.search(r'FISCAL RECEIPT|CREDIT NOTE', block_text, re.IGNORECASE):
                continue

            cu = re.search(
                r'CU\s*INVOICE\s*(?:NUMBER|NO)\s*[:\-]?\s*(?P<value>.+?)\s*(?:\r?\n|$)',
                block_text,
                re.IGNORECASE,
            )
            date = re.search(
                r'\bDATE\s*[:\-]?\s*(?P<value>\d{1,2}/\d{1,2}/\d{2,4})',
                block_text,
                re.IGNORECASE,
            )
            time = re.search(
                r'\bTIME\s*[:\-]?\s*(?P<value>\d{1,2}:\d{2}(?::\d{2})?)',
                block_text,
                re.IGNORECASE,
            )
            total = re.search(
                r'\bTOTAL(?:\s+AMOUNT)?\s*[:\-]?\s*(?P<value>[\d,]+\.?\d*)',
                block_text,
                re.IGNORECASE,
            )
            tax = re.search(
                r'\bTAX(?:\s+AMOUNT)?\s*[:\-]?\s*(?P<value>[\d,]+\.?\d*)',
                block_text,
                re.IGNORECASE,
            )
            entry_type = re.search(
                r'\b(FISCAL RECEIPT|CREDIT NOTE)\b',
                block_text,
                re.IGNORECASE,
            )
            if cu or date or time or total or tax or entry_type:
                rows.append([
                    cu.group('value').strip() if cu else '',
                    date.group('value').strip() if date else '',
                    time.group('value').strip() if time else '',
                    total.group('value').replace(' ', '').strip() if total else '',
                    tax.group('value').replace(' ', '').strip() if tax else '',
                    entry_type.group(1).upper() if entry_type else '',
                ])
        if rows:
            return [
                'CU Invoice Number', 'Date', 'Time',
                'Total', 'Tax Amount', 'Entry Type'
            ], rows
        return [], []

    try:
        if ext in ('.xlsx', '.xls'):
            wb = openpyxl.load_workbook(att.file.open('rb'), data_only=True)
            ws = wb.active
            rows = [tuple(cell if cell is not None else '' for cell in row) for row in ws.iter_rows(values_only=True)]
            headers = [str(cell) for cell in rows[0]] if rows else []
            return headers, [list(row) for row in rows[1:]] if len(rows) > 1 else []

        if ext == '.csv':
            with att.file.open('rb') as f:
                txt = f.read().decode('utf-8', errors='replace')
            reader = csv.reader(StringIO(txt))
            rows = [row for row in reader if any(cell.strip() for cell in row)]
            if rows:
                return [str(cell) for cell in rows[0]], [list(row) for row in rows[1:]]

        if ext == '.txt':
            with att.file.open('rb') as f:
                txt = f.read().decode('utf-8', errors='replace')
            headers, rows = _parse_txt_invoice_blocks(txt)
            if headers and rows:
                return headers, rows

            lines = [line.strip() for line in txt.splitlines() if line.strip()]
            if not lines:
                return [], []
            delimiter = '\t' if '\t' in lines[0] else ',' if ',' in lines[0] else None
            if delimiter:
                reader = csv.reader(lines, delimiter=delimiter)
                rows = [row for row in reader if any(cell.strip() for cell in row)]
                if rows:
                    return [str(cell) for cell in rows[0]], [list(row) for row in rows[1:]]

            # Fallback: try to capture key/value pairs into one row
            fields = {
                'CU Invoice Number': r'CU\s*INVOICE\s*NUMBER\s*[:\-]?\s*(.+)',
                'Date': r'\bDATE\s*[:\-]?\s*(.+)',
                'Time': r'\bTIME\s*[:\-]?\s*(.+)',
                'Total': r'\bTOTAL\s*[:\-]?\s*([\d,]+\.?\d*)',
                'Tax Amount': r'\bTAX(?:\s+AMOUNT)?\s*[:\-]?\s*([\d,]+\.?\d*)',
                'Entry Type': r'\b(FISCAL RECEIPT|CREDIT NOTE)\b',
                'Customer': r'\bCLIENT NAME\s*[:\-]?\s*(.+)',
            }
            row = []
            any_value = False
            for label, pattern in fields.items():
                match = re.search(pattern, txt, re.IGNORECASE)
                value = match.group(1).strip() if match else ''
                if value:
                    any_value = True
                row.append(value)
            if any_value:
                headers = list(fields.keys())
                return headers, [row]

            return [], []

        if ext == '.pdf':
            try:
                import pdfplumber
                with pdfplumber.open(att.file.open('rb')) as pdf:
                    for page in pdf.pages[:5]:
                        tables = page.extract_tables() or []
                        if tables:
                            for table in tables:
                                rows = [row for row in table if any(cell is not None and str(cell).strip() for cell in row)]
                                if rows:
                                    return [str(cell) for cell in rows[0]], [list(row) for row in rows[1:]]
            except Exception:
                pass
            return [], []
    except Exception as exc:
        logger.warning('Local extraction failed for %s: %s', att.filename, exc)
        return [], []


def _build_output_files_from_attachments(attachments):
    outputs = []
    for att in attachments:
        headers, rows = _extract_local_data_from_attachment(att)
        if headers and rows is not None:
            filename = Path(att.filename).stem + '_extracted.xlsx'
            buf = _build_xlsx(headers, rows, sheet_name=Path(filename).stem[:31])
            outputs.append({
                'filename': filename,
                'content': buf,
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
    return outputs


def _workflow_source_for_type(ext: str) -> str:
    return {
        'txt': 'ura_fiscal',
        'pdf': 'safaricom_bill',
        'xlsx': 'acon_sales',
        'xls': 'acon_sales',
        'csv': 'other',
    }.get(ext, 'other')


def _create_chat_batch(user, file_types: list[str]) -> UploadBatch:
    label = f"chat_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
    return UploadBatch.objects.create(
        uploaded_by=user,
        label=label,
        description="Chat attachment processing batch",
        status='processing',
    )


def _run_conversion_tools_for_attachments(file_attachments, user):
    """
    Attempt to ingest attachments through the UploadService pipeline.
    Returns a list of output file dicts if any conversions produced files.
    """
    outputs = []
    if not file_attachments:
        return outputs

    batch = _create_chat_batch(user, [att.file_type for att in file_attachments])
    for att in file_attachments:
        try:
            att.file.open('rb')
            data = att.file.read()
            from django.core.files.base import ContentFile as DjContentFile
            upload_obj = DjContentFile(data, name=att.filename)
            UploadService.ingest_file(batch, upload_obj)
        except Exception as exc:
            logger.warning('Tool conversion failed for %s: %s', att.filename, exc)

    batch.refresh_from_db()
    return outputs


# ─────────────────────────────────────────────────────────────────────────────
# ChatService
# ─────────────────────────────────────────────────────────────────────────────

class ChatService:

    @staticmethod
    def get_response(
        message: str,
        user,
        file_attachments=None,
        workflow_id: int | None = None,
        workflow_option: str | None = None,
        conversation_history: list[dict] | None = None,
    ) -> tuple[str, list]:
        """
        Returns (response_text, output_files).
        output_files is a list of {"filename", "content" (BytesIO), "content_type"}.
        """
        if not getattr(settings, "XAI_API_KEY", ""):
            return ChatService._offline_fallback(message), []

        has_files = bool(file_attachments)

        # File-processing requests go through the AI agent so they use the real
        # domain handlers (URA/Safaricom/ACON extraction, reconciliation, report
        # generation) instead of the unreliable "ask Grok to return JSON" path.
        if has_files:
            return ChatService._run_agent(
                message=message,
                user=user,
                file_attachments=file_attachments,
                workflow_id=workflow_id,
                workflow_option=workflow_option,
                conversation_history=conversation_history,
            )

        intent = _detect_intent(message, has_files=has_files)
        system_prompt = _INTENT_SYSTEM_PROMPTS.get(intent, _GENERAL_SYSTEM_PROMPT)

        if workflow_option == 'full_pipeline':
            system_prompt = (
                "You are a financial workflow assistant. Execute a full end-to-end processing pipeline "
                "for the attached files: parse, normalise, extract headers and rows, create Excel output, "
                "and prepare results for download."
            )
        elif workflow_id:
            try:
                workflow = Workflow.objects.get(pk=workflow_id)
                system_prompt = (
                    f"You are a financial workflow assistant. Apply the workflow '{workflow.name}' "
                    f"with steps: {', '.join(workflow.steps)}."
                )
            except Workflow.DoesNotExist:
                pass

        # Inject file content into the user message
        enriched = message
        if has_files:
            sections = []
            for att in file_attachments:
                try:
                    content = _read_file_to_text(att.file.open("rb"), att.filename)
                    sections.append(content)
                except Exception as exc:
                    logger.warning("Could not read %s: %s", att.filename, exc)
                    sections.append(f"[Could not read {att.filename}: {exc}]")
            enriched = (
                f"{message}\n\n"
                "=== ATTACHED FILE CONTENT ===\n"
                + "\n\n---\n\n".join(sections)
            )

        if intent == "status":
            enriched = f"{enriched}\n\n[Live batch data]\n{_get_status_context()}"

        try:
            raw = generate_chat_response(
                user_message=enriched,
                system_prompt=system_prompt,
                conversation_history=conversation_history,
            )
        except Exception as exc:
            logger.exception("Grok call failed for user %s", getattr(user, "username", "?"))
            return "I couldn't reach the AI service right now. Please try again.", []

        # For create_file intent try to parse JSON and build the xlsx
        output_files = []
        response_text = raw

        if intent == "create_file":
            parsed, xlsx_buf, filename, summary = ChatService._try_build_file(raw)
            if xlsx_buf:
                output_files.append({
                    "filename": filename,
                    "content":  xlsx_buf,
                    "content_type": (
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                })
                row_count = len(parsed.get("rows", []))
                col_count = len(parsed.get("headers", []))
                response_text = (
                    f"{summary}\n\n"
                    f"Extracted {row_count} rows across {col_count} columns. "
                    f"The file '{filename}' is attached — download it, review, "
                    "and re-upload for further processing if needed."
                )
            else:
                tool_outputs = _run_conversion_tools_for_attachments(file_attachments or [], user)
                if tool_outputs:
                    output_files = tool_outputs
                    response_text = (
                        "I could not parse a structured JSON response from AI, so I used the appropriate file conversion tools instead. "
                        "Please review the attached converted Excel file(s)."
                    )
                else:
                    output_files = _build_output_files_from_attachments(file_attachments or [])
                    if output_files:
                        response_text = (
                            "I could not parse a structured JSON response from AI, so I extracted the attached files locally instead. "
                            "Review the attached Excel file(s)."
                        )

        return response_text, output_files

    @staticmethod
    def _run_agent(
        *,
        message: str,
        user,
        file_attachments,
        workflow_id: int | None = None,
        workflow_option: str | None = None,
        conversation_history: list[dict] | None = None,
    ) -> tuple[str, list]:
        """
        Route a file-processing chat turn through the AI agent so it uses the
        real domain handlers (detect → extract → reconcile → report) instead of
        asking Grok to hand-write JSON.
        """
        from ai_engine.services import AIEngineService

        # 1. Ingest the uploaded attachments into a batch — this populates
        #    extracted_text and gives the agent UploadedFile ids to act on.
        batch = UploadService.create_batch(
            label=f"Chat upload ({len(file_attachments)} file(s))",
            user=user,
        )
        for att in file_attachments:
            try:
                att.file.open("rb")
                rec = UploadService.ingest_file(batch, att.file)
                try:
                    att.uploaded_file = rec
                    att.save(update_fields=["uploaded_file"])
                except Exception:
                    pass
            except Exception as exc:
                logger.warning(
                    "Chat agent could not ingest %s: %s",
                    getattr(att, "filename", "?"), exc,
                )
        batch.refresh_from_db()

        # 2. Optional workflow (constrains which tools the agent is offered).
        workflow = None
        if workflow_id:
            workflow = Workflow.objects.filter(pk=workflow_id).first()

        # 3. Run the agent tool loop.
        try:
            response_text, job_id = AIEngineService.handle_chat_message(
                user=user,
                message=message,
                batch=batch,
                workflow=workflow,
                conversation_history=conversation_history,
            )
        except Exception as exc:
            logger.exception("Chat agent run failed: %s", exc)
            return f"Sorry, I hit an error while processing the file: {exc}", []

        # 4. Surface any files the tools produced as downloadable attachments.
        output_files = ChatService._collect_job_output_files(job_id)
        if output_files and "attach" not in response_text.lower():
            names = ", ".join(f["filename"] for f in output_files)
            response_text = f"{response_text}\n\nGenerated file(s) attached: {names}"
        return response_text, output_files

    @staticmethod
    def _collect_job_output_files(job_id) -> list:
        """Read any .xlsx/.csv/.pdf files referenced in a job's successful tool results."""
        from pathlib import Path
        from io import BytesIO
        from tools.models import ToolCall

        mime = {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "csv":  "text/csv",
            "pdf":  "application/pdf",
        }
        files, seen = [], set()
        if not job_id:
            return files
        for tc in ToolCall.objects.filter(job_id=job_id, status="success"):
            for value in (tc.result or {}).values():
                if not isinstance(value, str) or "." not in value:
                    continue
                ext = value.rsplit(".", 1)[-1].lower()
                if ext not in mime or value in seen:
                    continue
                p = Path(value)
                if p.exists() and p.is_file():
                    seen.add(value)
                    files.append({
                        "filename":     p.name,
                        "content":      BytesIO(p.read_bytes()),
                        "content_type": mime[ext],
                    })
        return files

    @staticmethod
    def _try_build_file(raw: str) -> tuple:
        """Parse Grok JSON → (dict, BytesIO|None, filename, summary)."""
        text = raw.strip()
        # Strip markdown fences
        if text.startswith("```"):
            text = text.split("```", 2)[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.rsplit("```", 1)[0].strip()

        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            logger.warning("create_file: could not parse Grok JSON")
            return {}, None, "output.xlsx", raw

        headers  = data.get("headers", [])
        rows     = data.get("rows", [])
        filename = data.get("filename", "extracted_data.xlsx")
        summary  = data.get("summary", "Data extracted from the uploaded file.")

        if not headers or not rows:
            return data, None, filename, summary

        if not filename.endswith(".xlsx"):
            filename = filename.rsplit(".", 1)[0] + ".xlsx"

        xlsx_buf = _build_xlsx(headers, rows, sheet_name=Path(filename).stem[:31])
        return data, xlsx_buf, filename, summary

    @staticmethod
    def _offline_fallback(message: str) -> str:
        intent = _detect_intent(message)
        fallbacks = {
            "help":        "I can help with uploads, parsing, reconciliation, AI analysis, and reports.",
            "upload":      "Create a batch (POST /api/uploads/batches/), then upload to POST /api/uploads/batches/<id>/files/.",
            "process":     "Files parse automatically after upload. Check at GET /api/uploads/batches/<id>/summary/.",
            "reconcile":   "POST /api/tools/reconciliations/ with batch and reconcile_type.",
            "report":      "POST /api/tools/reports/ then download at GET /api/tools/reports/<id>/download/.",
            "status":      "GET /api/uploads/batches/?status=pending",
            "ai_analysis": "POST /api/ai/jobs/ with task_type: flag_anomalies, summarise_batch, or explain_variance.",
            "create_file": "File processing requires XAI_API_KEY to be set in your .env.",
        }
        return fallbacks.get(intent, "I can help with uploads, processing, reconciliation, and reports.")
