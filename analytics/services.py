# analytics/services.py
"""
ReportBuildService — turns a Report row's (report_type, parameters) into a
real downloadable file and writes it onto Report.file. Called from
analytics.tasks.generate_report_task, kept out of the view/task so the
actual "what data goes in this report" logic lives in exactly one place.
"""
import csv
import io
import logging

from django.core.files.base import ContentFile
from django.utils import timezone

from .models import Report

logger = logging.getLogger(__name__)


def _safe_import():
    from uploads.models import UploadedFile
    from ai_engine.models import AIAnalysisJob, AIInsight
    return UploadedFile, AIAnalysisJob, AIInsight


class ReportBuildService:

    @staticmethod
    def build(report: Report) -> None:
        headers, rows = ReportBuildService._collect_rows(report)

        if report.format == "xlsx":
            content = ReportBuildService._build_xlsx(report.name, headers, rows)
            ext = "xlsx"
        elif report.format == "csv":
            content = ReportBuildService._build_csv(headers, rows)
            ext = "csv"
        else:
            content = ReportBuildService._build_pdf(report.name, headers, rows)
            ext = "pdf"

        safe_name = "".join(c for c in report.name if c.isalnum() or c in " _-")[:100] or "report"
        report.file.save(f"{safe_name}.{ext}", ContentFile(content), save=False)
        report.file_size = len(content)
        report.status = "ready"
        report.generated_at = timezone.now()
        report.save(update_fields=["file", "file_size", "status", "generated_at"])

    # ── Row collection ──────────────────────────────────────────────────────

    @staticmethod
    def _collect_rows(report: Report) -> tuple[list[str], list[list]]:
        UploadedFile, AIAnalysisJob, AIInsight = _safe_import()
        org = report.organization
        params = report.parameters or {}
        batch_id = params.get("batch_id")
        date_from = params.get("date_from")
        date_to = params.get("date_to")

        if report.report_type == "reconciliation":
            qs = AIInsight.objects.filter(job__organization=org).select_related("job")
            if batch_id:
                qs = qs.filter(job__batch_id=batch_id)
            if date_from:
                qs = qs.filter(created_at__date__gte=date_from)
            if date_to:
                qs = qs.filter(created_at__date__lte=date_to)
            headers = ["Job ID", "Type", "Severity", "Title", "Detail", "Created At"]
            rows = [
                [
                    ins.job_id, ins.get_insight_type_display(), ins.severity,
                    ins.title, ins.detail, ins.created_at.isoformat(),
                ]
                for ins in qs.order_by("-created_at")
            ]
            return headers, rows

        if report.report_type == "analytics":
            from datetime import timedelta
            now = timezone.now()
            since = timezone.datetime.fromisoformat(str(date_from)) if date_from else now - timedelta(days=30)
            until = timezone.datetime.fromisoformat(str(date_to)) if date_to else now
            if timezone.is_naive(since):
                since = timezone.make_aware(since)
            if timezone.is_naive(until):
                until = timezone.make_aware(until)

            headers = ["Date", "Files Uploaded", "Files Parsed", "Jobs Run", "Jobs Done"]
            rows = []
            days = max(1, (until - since).days)
            for i in range(min(days, 366)):
                day_start = (since + timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
                day_end = day_start + timedelta(days=1)
                uploaded = UploadedFile.objects.filter(
                    batch__organization=org, uploaded_at__gte=day_start, uploaded_at__lt=day_end,
                ).count()
                parsed = UploadedFile.objects.filter(
                    batch__organization=org, uploaded_at__gte=day_start, uploaded_at__lt=day_end,
                    parse_status="parsed",
                ).count()
                jobs_run = AIAnalysisJob.objects.filter(
                    organization=org, created_at__gte=day_start, created_at__lt=day_end,
                ).count()
                jobs_done = AIAnalysisJob.objects.filter(
                    organization=org, created_at__gte=day_start, created_at__lt=day_end, status="done",
                ).count()
                rows.append([day_start.strftime("%Y-%m-%d"), uploaded, parsed, jobs_run, jobs_done])
            return headers, rows

        # "billing" and "custom" — a listing of uploaded files, the closest
        # generic thing every org has regardless of report_type nuance.
        qs = UploadedFile.objects.filter(batch__organization=org).select_related("batch")
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        if date_from:
            qs = qs.filter(uploaded_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(uploaded_at__date__lte=date_to)
        headers = ["Filename", "Batch", "Detected Type", "Parse Status", "Size (bytes)", "Uploaded At"]
        rows = [
            [
                uf.original_filename, uf.batch.label, uf.detected_type or "", uf.parse_status,
                uf.file_size_bytes, uf.uploaded_at.isoformat(),
            ]
            for uf in qs.order_by("-uploaded_at")
        ]
        return headers, rows

    # ── File builders ────────────────────────────────────────────────────────

    @staticmethod
    def _build_xlsx(title: str, headers: list[str], rows: list[list]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D3B8E", end_color="0D3B8E", fill_type="solid")
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value if isinstance(value, (int, float, str)) else str(value))

        for col_idx, header in enumerate(headers, start=1):
            if col_idx > 26:
                break
            ws.column_dimensions[chr(64 + col_idx)].width = max(12, len(str(header)) + 2)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _build_csv(headers: list[str], rows: list[list]) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return buf.getvalue().encode("utf-8")

    @staticmethod
    def _build_pdf(title: str, headers: list[str], rows: list[list]) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()

        elements = [
            Paragraph(title, styles["Title"]),
            Paragraph(f"Generated at {timezone.now().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
            Spacer(1, 0.25 * inch),
        ]

        table_data = [headers] + [[str(v) for v in row] for row in rows] if rows else [headers]
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D3B8E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F7")]),
        ]))
        elements.append(table)

        doc.build(elements)
        return buf.getvalue()
