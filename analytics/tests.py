from rest_framework.test import APIClient
from django.test import TestCase

from organizations.models import Organization
from users.models import User
from uploads.services import UploadService
from analytics.models import Report


class DashboardProcessingViewTests(TestCase):
    """
    DashboardProcessingView's response keys must match the frontend's
    ProcessingDataPoint shape ("scrutinized"/"flagged") — the frontend was
    rebranded from "invoices"/"reconciled" terminology but this view wasn't
    updated to match, so the Weekly Scrutiny & Flag Trend chart silently
    rendered nothing against a live backend.
    """

    def setUp(self):
        self.user = User.objects.create_user(
            email="dash@example.com", password="pw12345!",
            first_name="D", last_name="B",
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_processing_response_uses_scrutinized_and_flagged_keys(self):
        response = self.client.get("/api/dashboard/processing/?period=week")
        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertGreater(len(data), 0)
        for point in data:
            self.assertIn("name", point)
            self.assertIn("scrutinized", point)
            self.assertIn("flagged", point)
            self.assertNotIn("invoices", point)
            self.assertNotIn("reconciled", point)


class DashboardOrgIsolationTests(TestCase):
    """
    Every dashboard/analytics view queried UploadedFile/AIAnalysisJob/Report
    platform-wide before Phase 3, with zero ownership filtering — any
    authenticated user could see every other user's stats and recent
    filenames. Verify aggregate counts and the recent-activity feed only
    ever reflect this org's data, and that Report visibility is now
    org-shared (not just requester-private).
    """

    def setUp(self):
        self.org_a = Organization.objects.create(name="Org A")
        self.org_b = Organization.objects.create(name="Org B")
        self.user_a = User.objects.create_user(
            email="usera@example.com", password="pw12345!",
            first_name="U", last_name="A", organization=self.org_a,
        )
        self.teammate_a = User.objects.create_user(
            email="teammatea@example.com", password="pw12345!",
            first_name="T", last_name="A", organization=self.org_a,
        )
        self.user_b = User.objects.create_user(
            email="userb@example.com", password="pw12345!",
            first_name="U", last_name="B", organization=self.org_b,
        )
        self.client = APIClient()

    def _upload_for(self, user, filename="doc.txt"):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from unittest.mock import patch
        batch = UploadService.create_batch(label=f"{user.email} batch", user=user)
        upload = SimpleUploadedFile(filename, b"content", content_type="text/plain")
        with patch("uploads.services._extract_text", return_value="hello"):
            return UploadService.ingest_file(batch, upload)

    def test_dashboard_stats_only_counts_this_org(self):
        self._upload_for(self.user_a)
        self._upload_for(self.user_b)
        self._upload_for(self.user_b)

        self.client.force_authenticate(user=self.user_a)
        resp = self.client.get("/api/dashboard/stats/")
        self.assertEqual(resp.status_code, 200)
        invoices_stat = next(s for s in resp.json()["stats"] if s["title"] == "Invoices Processed")
        self.assertEqual(invoices_stat["value"], "1")

    def test_recent_activity_does_not_leak_other_org_filenames(self):
        self._upload_for(self.user_a, filename="org_a_only.txt")
        self._upload_for(self.user_b, filename="org_b_secret.txt")

        self.client.force_authenticate(user=self.user_a)
        resp = self.client.get("/api/dashboard/recent-activity/")
        self.assertEqual(resp.status_code, 200)
        titles = " ".join(a["title"] for a in resp.json()["activities"])
        self.assertIn("org_a_only.txt", titles)
        self.assertNotIn("org_b_secret.txt", titles)

    def test_report_is_org_shared_not_requester_private(self):
        """Reports flip from user-private to org-shared in Phase 3 — a
        teammate who didn't request the report can still see/download it."""
        report = Report.objects.create(
            name="Org A Report", report_type="custom", format="csv",
            requested_by=self.user_a, organization=self.org_a, status="ready",
        )

        self.client.force_authenticate(user=self.teammate_a)
        resp = self.client.get("/api/reports/")
        self.assertEqual(resp.status_code, 200)
        ids = [r["id"] for r in resp.json()["reports"]]
        self.assertIn(report.pk, ids)

        self.client.force_authenticate(user=self.user_b)
        resp = self.client.get("/api/reports/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["reports"], [])


class ReportGenerationTests(TestCase):
    """
    ReportGenerateView used to force status="ready" immediately with no file
    ever written — every download hit "Report file not available.". Now it
    dispatches a real Celery task (generate_report_task) that builds a real
    file via ReportBuildService.
    """

    def setUp(self):
        self.org = Organization.objects.create(name="Org")
        self.user = User.objects.create_user(
            email="reports@example.com", password="pw12345!",
            first_name="R", last_name="P", organization=self.org,
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def _upload_for(self, user, filename="doc.txt"):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from unittest.mock import patch
        batch = UploadService.create_batch(label=f"{filename} batch", user=user)
        upload = SimpleUploadedFile(filename, b"content", content_type="text/plain")
        with patch("uploads.services._extract_text", return_value="hello"):
            return UploadService.ingest_file(batch, upload)

    def test_generate_dispatches_task_and_stays_generating(self):
        from unittest.mock import patch
        with patch("config.dispatch.dispatch", return_value="queued") as mock_dispatch:
            resp = self.client.post("/api/reports/generate/", {
                "type": "custom", "format": "csv", "parameters": {},
            }, format="json")
        self.assertEqual(resp.status_code, 202)
        self.assertTrue(mock_dispatch.called)
        report = Report.objects.get(pk=resp.json()["reportId"])
        self.assertEqual(report.status, "generating")
        self.assertFalse(report.file)

    def test_generate_marks_error_when_broker_unavailable(self):
        """A dispatch failure must mark the report terminally errored."""
        from unittest.mock import patch

        with patch("config.dispatch.dispatch", return_value=None):
            resp = self.client.post("/api/reports/generate/", {
                "type": "custom", "format": "csv", "parameters": {},
            }, format="json")
        self.assertEqual(resp.status_code, 202)
        report = Report.objects.get(pk=resp.json()["reportId"])
        self.assertEqual(report.status, "error")

    def test_generate_report_task_produces_downloadable_xlsx_with_real_content(self):
        from analytics.tasks import generate_report_task

        self._upload_for(self.user, "invoice1.txt")
        self._upload_for(self.user, "invoice2.txt")

        report = Report.objects.create(
            name="Test Billing Report", report_type="billing", format="xlsx",
            requested_by=self.user, organization=self.org, status="generating",
        )
        generate_report_task(report.id)

        report.refresh_from_db()
        self.assertEqual(report.status, "ready")
        self.assertTrue(report.file)
        self.assertGreater(report.file_size, 0)

        resp = self.client.get(f"/api/reports/{report.pk}/download/")
        self.assertEqual(resp.status_code, 200)
        content = b"".join(resp.streaming_content) if resp.streaming else resp.content

        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "Filename")
        # Header row + 2 uploaded files.
        self.assertEqual(ws.max_row, 3)

    def test_generate_report_task_pdf_and_csv_formats_also_work(self):
        from analytics.tasks import generate_report_task

        self._upload_for(self.user)

        pdf_report = Report.objects.create(
            name="PDF Report", report_type="billing", format="pdf",
            requested_by=self.user, organization=self.org, status="generating",
        )
        generate_report_task(pdf_report.id)
        pdf_report.refresh_from_db()
        self.assertEqual(pdf_report.status, "ready")
        self.assertTrue(pdf_report.file.read().startswith(b"%PDF"))

        csv_report = Report.objects.create(
            name="CSV Report", report_type="billing", format="csv",
            requested_by=self.user, organization=self.org, status="generating",
        )
        generate_report_task(csv_report.id)
        csv_report.refresh_from_db()
        self.assertEqual(csv_report.status, "ready")

        import csv as csv_module
        from io import StringIO
        csv_report.file.open("r")
        text = csv_report.file.read()
        csv_report.file.close()
        if isinstance(text, bytes):
            text = text.decode("utf-8")
        rows = list(csv_module.reader(StringIO(text)))
        self.assertEqual(rows[0][0], "Filename")
        self.assertEqual(len(rows), 2)  # header + 1 uploaded file

    def test_generate_report_task_marks_error_on_exception(self):
        from unittest.mock import patch
        from analytics.tasks import generate_report_task

        report = Report.objects.create(
            name="Broken Report", report_type="billing", format="xlsx",
            requested_by=self.user, organization=self.org, status="generating",
        )
        with patch("analytics.services.ReportBuildService.build", side_effect=RuntimeError("boom")):
            with self.assertRaises(RuntimeError):
                generate_report_task(report.id)

        report.refresh_from_db()
        self.assertEqual(report.status, "error")
        self.assertIn("boom", report.error_message)

    def test_report_download_400s_when_not_ready(self):
        report = Report.objects.create(
            name="Still Generating", report_type="billing", format="xlsx",
            requested_by=self.user, organization=self.org, status="generating",
        )
        resp = self.client.get(f"/api/reports/{report.pk}/download/")
        self.assertEqual(resp.status_code, 400)
